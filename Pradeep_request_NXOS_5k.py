import openpyxl
from netmiko import ConnectHandler
import re
import funtions_jose
from openpyxl import Workbook, load_workbook
from openpyxl.styles.alignment import Alignment

print('==> Pradeep Request Script <==')

if __name__ == '__main__':
    print('=> Pradeep Request for 5k info..')

    v_boldFont = openpyxl.styles.Font(bold=True)
    v_centerAlignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrapText=True)

    # creating the Excel file
    # wb = load_workbook('dev/Pradeep.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.title = 'FiveK info'  # creating sheet title
    ws.append(['Hostname', 'Description', 'Device SN', 'NXOS version', 'Fex Type', 'Fex SN'])

    # loop for all devices
    with open("dev/fiveK.txt", 'r') as hostsfile:
        for line in hostsfile:
            hostline = line.strip()
            isIP = line

            JC = funtions_jose.if_credential_connection(isIP)
            net_connect = ConnectHandler(**JC)
            net_connect.enable()

            # printing the IP
            print('=> This is the current device', isIP)

            # Getting version of the box
            version = net_connect.send_command('sh version', read_timeout=703)
            version_detail = re.search(r'sys\w+:\s+\w+\s(.*)', version).group(1)
            # print(version_detail)

            # Getting hostname
            hostname = net_connect.send_command('sh run', read_timeout=803)
            hostname_detail = re.search(r'hostn\w+\s(.*)', hostname).group(1)
            # print(hostname_detail)

            # getting inventory
            inv = net_connect.send_command('sh inventory', read_timeout=803)
            inv_detail = re.search(r'Chas\w+..\sDE\w+.\s(\S+\s.*)[\r\n]+\S+\s\S+\s+.\s\S+\s\S+\s.\s\S+\s(.*)', inv)
            # group 1 = description
            # group 2 = SN
            # print(inv_detail.group(1))
            # print(inv_detail.group(2))

            # getting flex detail
            fex = net_connect.send_command('sh fex deta', read_timeout=803)
            if 'state' in fex:
                fex_SN = re.search(r'Ex\w+\sSe\S+\s(.*)', fex).group(1)
                fex_device_type = re.search(r'Ex\w+\sMo\S+\s(\S+)', fex).group(1)
                # removing tail from fex_device_type
                fex_device = fex_device_type.rstrip(',')
                # print(fex_SN)
                # print(fex_device)
            else:
                fex_SN = 'none'
                fex_device = 'none'

            # Excel work
            # loading a excell file
            # wb = load_workbook('example.xlsx')
            # ws = wb.active
            # print(ws['A1'].value) # print a value from a cell
            # ws['A2'].value = 'value' # to set a value
            # print(wb.sheetnames) # to see all the sheets inside the Excel
            # print(wb['Sheet1']) to access sheet1
            # wb.create_sheet('Test') # creating a sheet called test
            # wb = Workbook() # create a new Excel file
            # wb.save('example.xlsx') # to save the file, always required to see the changes

            # filling out the Exel
            ws.append([hostname_detail, inv_detail.group(1), inv_detail.group(2), version_detail, fex_device, fex_SN])

    # Bold and align to the Center
    for cell in ws[1]:
        cell.font = v_boldFont
        cell.alignment = v_centerAlignment

    # saving the Excel
    wb.save('Pradeep.xlsx')
