import openpyxl
from netmiko import ConnectHandler
import re
import funtions_jose
from openpyxl import Workbook, load_workbook
from openpyxl.styles.alignment import Alignment

print('==> Script to check Default AP names <==')

if __name__ == '__main__':

    v_boldFont = openpyxl.styles.Font(bold=True)
    v_centerAlignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrapText=True)

    # creating the Excel file
    wb = load_workbook('excel_output/WLC-Default_Name.xlsx')
    # wb = Workbook()
    ws = wb.active
    ws.title = 'WLC_Default_Names'  # creating sheet title
    ws.append(['WLC', 'AP_Name', 'AP_Model', 'MAC_ADD', 'IP_ADD', 'Clients_Connected'])

    # loop for all devices AireOS
    with open("wlc-airOS.txt", 'r') as hostsfile:
        for line in hostsfile:
            hostline = line.strip()
            isIP = line

            JC = funtions_jose.connect_wlc(isIP)
            net_connect = ConnectHandler(**JC)
            net_connect.enable()

            # printing the IP
            print('=> This is the current device', isIP)

            # Getting hostname
            hostname = net_connect.send_command('show sysinfo', read_timeout=803)
            hostname_detail = re.search(r'Sys\w+\sName\S+\s(.*)', hostname).group(1)
            # print(hostname_detail)

            # loop for ap_summary
            ap_summ = net_connect.send_command("show ap summary", read_timeout=803)
            ap_filter = re.findall(
                r'(AP[0-9A-Fa-f]{4}\.[0-9A-Fa-f]{4}\.[0-9A-Fa-f]{4})\s+\d+\s+(\S+)\s+(\S+)\s+\S+\s\S+\s+\w+\s+(\S+)\s+(\d+)',
                ap_summ)

            # getting the total of Default AP name
            if len(ap_filter) == 1:
                print('=> The total of Default AP names is', len(ap_filter))
            elif len(ap_filter) > 1:
                print('=> The total of Default AP names are', len(ap_filter))

            for item in ap_filter:
                ap_name = item[0]
                ap_model = item[1]
                ap_mac_add = item[2]
                ap_ip = item[3]
                ap_client_count = item[4]

                # filling out the Exel
                ws.append([hostname_detail, ap_name, ap_model, ap_mac_add, ap_ip, ap_client_count])
                ws.alignment = v_centerAlignment

            # Excel work
            # loading a excell file
            # wb = load_workbook('example.xlsx')
            # ws = wb.active
            # print(ws['A1'].value) # print a value from a cell
            # ws['A2'].value = 'value' # to set a value
            # print(wb.sheetnames) # to see all the sheets inside the Excel
            # print(wb['Sheet1']) to access sheet1
            # wb.create_sheet('Test') # creating a sheet called test
            # wb = Workbook() # create a new Excel file
            # wb.save('example.xlsx') # to save the file, always required to see the changes

    # loop for all devices IOS
    with open("Inventory/wlc-IOS.txt", 'r') as hostsfile:
        for line in hostsfile:
            hostline = line.strip()
            isIP = line

            JC = funtions_jose.if_credential_connection(isIP)
            net_connect = ConnectHandler(**JC)
            net_connect.enable()

            # printing the IP
            print('=> This is the current device', isIP)

            # Getting hostname
            hostname = funtions_jose.get_hostname_only(net_connect)

            # loop for ap_summary
            ap_summ = net_connect.send_command("show ap summary", read_timeout=903)
            ap_filter = re.findall(
                r'(AP[0-9A-Fa-f]{4}\.[0-9A-Fa-f]{4}\.[0-9A-Fa-f]{4})\s+\d+\s+(\S+)\s+(\S+)\s+\S+\s+\S+\s\S+\s+\w+\s+(\S+)', ap_summ)

            # getting the total of Default AP name
            if len(ap_filter) == 1:
                print('=> The total of Default AP names is', len(ap_filter))
            elif len(ap_filter) > 1:
                print('=> The total of Default AP names are', len(ap_filter))

            for item in ap_filter:
                ap_name = item[0]
                ap_model = item[1]
                ap_mac_add = item[2]
                ap_ip = item[3]

                # filling out the Exel
                ws.append([hostname, ap_name, ap_model, ap_mac_add, ap_ip, '-'])

    # Bold and align to the Center
    for cell in ws[1]:
        cell.font = v_boldFont
        cell.alignment = v_centerAlignment

    # saving the Excel
    wb.save('excel_output/WLC-Default_Name.xlsx')


