from netmiko import ConnectHandler
import getpass
import re
import funtions_jose
from openpyxl import Workbook, load_workbook
from openpyxl.styles.alignment import Alignment

print('==> script to get vlan network info <==')

if __name__ == '__main__':
    # XP
    FZs = []

    # creating the Excel file
    wb = load_workbook('Vlan21.xlsx')
    # wb = Workbook()
    ws = wb.active
    ws.title = 'Vlan21-Campus'  # creating sheet title
    ws.append(['Hostname', 'Network'])

    # loop for all WLCs
    with open("FZs.txt", 'r') as hostsfile:
        for line in hostsfile:
            hostline = line.strip()
            isIP = line

            print('==> Connecting to FZ dist switch =>' + " " + str(line))
            JC = funtions_jose.if_credential_connection(line)
            net_connect = ConnectHandler(**JC)
            net_connect.enable()

            # getting Hostname
            hostname_detail = net_connect.send_command('sh run | inc hostname')
            hostname = re.search(r'hostn\w+\s(.*)', hostname_detail).group(1)
            print(hostname)

            vl_21 = net_connect.send_command('sh run int vlan 21')
            # getting IP add
            ip = re.search(r'ip\saddr\w+\s(.*)', vl_21)
            only_ip_no_mask = re.search(r'^(\d+\.\d+\.\d+\.\d+)', ip.group(1))

            if '255.255.255' in ip.group(1):
                ip_cidr = re.search(r'\d{1,3}$', ip.group(1))
                if '128' in ip_cidr.group():
                    ipv4 = only_ip_no_mask.group(1) + '/25'
                    # filling out the Exel
                    ws.append([hostname_detail, ipv4])
                if '192' in ip_cidr.group():
                    ipv4 = only_ip_no_mask.group(1) + '/26'
                    # filling out the Exel
                    ws.append([hostname_detail, ipv4])
                if '224' in ip_cidr.group():
                    ipv4 = only_ip_no_mask.group(1) + '/27'
                    # filling out the Exel
                    ws.append([hostname_detail, ipv4])
                if '240' in ip_cidr.group():
                    ipv4 = only_ip_no_mask.group(1) + '/28'
                    # filling out the Exel
                    ws.append([hostname_detail, ipv4])
                if '248' in ip_cidr.group():
                    ipv4 = only_ip_no_mask.group(1) + '/29'
                    # filling out the Exel
                    ws.append([hostname_detail, ipv4])
                if '252' in ip_cidr.group():
                    ipv4 = only_ip_no_mask.group(1) + '/30'
                    # filling out the Exel
                    ws.append([hostname_detail, ipv4])
                if '254' in ip_cidr.group():
                    ipv4 = only_ip_no_mask.group(1) + '/31'
                    # filling out the Exel
                    ws.append([hostname_detail, ipv4])
                if '255' in ip_cidr.group():
                    ipv4 = only_ip_no_mask.group(1) + '/32'
                    # filling out the Exel
                    ws.append([hostname_detail, ipv4])

    # saving the Excel
    wb.save('Vlan21.xlsx')
