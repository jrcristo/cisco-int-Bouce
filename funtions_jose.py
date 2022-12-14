from __future__ import print_function
import requests
import json
import datetime
import getpass
from builtins import print
from tenacity import sleep
from snmp_helper import snmp_get_oid, snmp_extract
from netmiko import ConnectHandler
import re
import os
import smtplib
from email.mime.multipart import MIMEMultipart
import openpyxl
from email.mime.text import MIMEText
from openpyxl import Workbook, load_workbook
from openpyxl.styles.alignment import Alignment

yes_option = ['yes', 'y']
no_option = ['no', 'n']


def get_ios_logs_by_interface(interface, net_connect):
    print('*-----*.*-----*.*-----*.*-----*.*-----*.')
    logs = net_connect.send_command('sh log | inc ' + interface, read_timeout=603)
    print(logs)


def cisco_prime_api_results(ap_name, pi_ip):
    requests.packages.urllib3.disable_warnings()

    base_uri = 'https://' + pi_ip + '/webacs/api/v4/'
    user = 'xopsapi'
    password = 'Xops@123'
    # rest_path = '/data/InventoryDetails'
    # rest_path = 'data/AccessPoints?name=eq("Cabana_11")'
    rest_path = 'data/AccessPoints?name=eq("' + ap_name + '")'
    url = base_uri + rest_path

    headers = {'Accept': 'application/json'}
    response = requests.get(url, headers=headers, auth=(user, password), verify=False)
    # print(response.text)
    # print(response.status_code)98745
    acdResp = json.loads(response.text)
    # print(acdResp)

    if acdResp['queryResponse']['@count'] == 1:
        print('=> Device Found it, getting details')

        # capturing device ID
        dev_id = acdResp['queryResponse']['entityId'][0]['$']
        # print('id =', dev_id)

        # showing device details
        # device = 'data/Devices/' + dev_id
        device = 'data/AccessPointDetails/' + dev_id
        url_dev = base_uri + device
        dev_details = requests.get(url_dev, headers=headers, auth=(user, password), verify=False)
        cdpData = json.loads(dev_details.text)
        cdpNeighList = cdpData['queryResponse']['entity'][0]['accessPointDetailsDTO']['cdpNeighbors']['cdpNeighbor']
        i = 0
        for neigh in cdpNeighList:
            i += 1
            # print("Neighbor: ", i)
            print('*-----*.*-----*.*-----*.*-----*.*-----*.')
            print("=> The last known CDP Neighbor name was:", neigh['neighborName'])
            print("=> The last known CDP Neighbor port was:", neigh['neighborPort'])
            print('=> The Last known IP neighbor address was:', neigh['neighborIpAddress']['address'])
            print('=> Switch type is:', neigh['platform'])

    else:
        print('=> Device not found it, exiting')
        exit(0)


def cisco_prime_api_results_devices(tl_name, pi_ip, domain, net_connect):
    requests.packages.urllib3.disable_warnings()

    base_uri = 'https://' + pi_ip + '/webacs/api/v4/'
    user = 'xopsapi'
    password = 'Xops@123'
    # rest_path = '/data/InventoryDetails'
    # rest_path = 'data/AccessPoints?name=eq("Cabana_11")'
    rest_path = 'data/Devices?deviceName=eq("' + tl_name + '")'
    url = base_uri + rest_path
    # print(url)

    headers = {'Accept': 'application/json'}
    response = requests.get(url, headers=headers, auth=(user, password), verify=False)
    # print(response.text)
    # print(response.status_code)98745
    acdResp = json.loads(response.text)
    # print(acdResp)

    if acdResp['queryResponse']['@count'] == 1:
        print('=> Device Found it, getting details')

        # capturing device ID
        dev_id = acdResp['queryResponse']['entityId'][0]['$']
        # print('id =', dev_id)

        # showing device details
        # device = 'data/Devices/' + dev_id
        device = 'data/InventoryDetails/' + dev_id
        url_dev = base_uri + device
        dev_details = requests.get(url_dev, headers=headers, auth=(user, password), verify=False)
        cdpData = json.loads(dev_details.text)
        # printing device Info
        print("This TL name is:", cdpData['queryResponse']['entity'][0]['inventoryDetailsDTO']['summary']['deviceName'])
        print("This TL last IP was:",
              cdpData['queryResponse']['entity'][0]['inventoryDetailsDTO']['summary']['ipAddress'])
        print("This TL model is:",
              cdpData['queryResponse']['entity'][0]['inventoryDetailsDTO']['summary']['deviceType'])
        print("This TL location is:",
              cdpData['queryResponse']['entity'][0]['inventoryDetailsDTO']['summary']['location'])
        print('*-----*.*-----*.*-----*.*-----*.*-----*.')
        print('==> CDP NEIGHBOR INFO <==')
        cdpNeighList = cdpData['queryResponse']['entity'][0]['inventoryDetailsDTO']['cdpNeighbors']['cdpNeighbor']
        i = 0
        for neigh in cdpNeighList:
            i += 1
            if 'GigabitEthernet0/9' in neigh['nearEndInterface']:
                print('*-----*.*-----*.*-----*.*-----*.*-----*.')
                # removing long tail or domain from the name, just keeping the name
                name = neigh['neighborDeviceName']
                new_name = re.search(r'(^[^.]+)', name)
                print(" -> The last known CDP Neighbor name was:", new_name.group() + domain)
                print(" -> The last known CDP Neighbor port was:", neigh['farEndInterface'])
                # print('=> The Last known IP neighbor address was:', neigh['neighborIpAddress']['address'])
                print(' -> The CDP device type is:', neigh['neighborDevicePlatformType'])

                # Validation for CB, IDF name does not contain cb in the FQDN
                if 'PCLCB' in new_name.group() and 'ID' in new_name.group():
                    domain = '.cruises.princess.com'
                # Validation for DI, IDF name does not contain cb in the FQDN
                if 'PCLDI' in new_name.group() and 'ID' in new_name.group() or 'LIDF' in new_name.group():
                    # Validation for IP, IDF name does not contain cb in the FQDN
                    domain = '.cruises.princess.com'

                # getting CDP neighbor IP
                neig_ip = cisco_prime_api_results_devices_IP(new_name.group() + domain, pi_ip)

                print(' -> The CDP Neigh IP is:', neig_ip)
                if '10.28' in neig_ip:
                    print("==>*** Neighbor is a Tech Locker, exiting ***<==")
                    exit(0)
                else:
                    pass
                # connecting to IDF and checking port status
                print(
                    '==> Connecting to IDF ' + new_name.group() + ' and checking the port ' + neigh['farEndInterface'])
                JC = if_credential_connection(neig_ip)
                net_connect = ConnectHandler(**JC)
                net_connect.enable()

                # checking cdp neigh
                output = net_connect.send_command('sh cdp ne ' + neigh['farEndInterface'], read_timeout=603)
                if 'Total cdp entries displayed : 0' in output:
                    print(' -> No CDP neighbor detected, checking power inline')
                    power_inline = net_connect.send_command('sh power inline ' + neigh['farEndInterface'],
                                                            read_timeout=603)
                    power_details = re.search(r'Gi\S+\s+\S+\s+(\w+)\s+(\S+)\s+(\S+)', power_inline)
                    print(' -> Operational PoE status is:', power_details.group(1))
                    print(
                        ' -> Max Power assigned on port ' + neigh['farEndInterface'] + ' is: ' + power_details.group(2))
                    print(' -> Device id:', power_details.group(3), '\n')
                    if 'n/a' in power_details.group(3) and 'off' in power_details.group(1):
                        print('==> Please, create a ticket and call the ITO. Port ' + neigh[
                            'farEndInterface'] + " " + 'is down <==')
                        exit(0)
                    else:
                        print('==> Testing Cable <==')
                        cable_test = net_connect.send_command(
                            'test cable-diagnostics tdr interface ' + neigh['farEndInterface'], read_timeout=603)
                        if 'TDR test started on interface' in cable_test:
                            print(' -> TDR test started')
                        sleep(11)
                        cable_test_result = net_connect.send_command(
                            'show cable-diagnostics tdr interface ' + neigh['farEndInterface'], read_timeout=603)
                        print(cable_test_result, '\n')
                        restart_int = input(
                            "==> do you want to reboot the interface facing the TL?, (Y) to continue (N) to cancel:").lower()
                        if restart_int in yes_option:
                            config_commands = ['int ' + neigh['farEndInterface'], 'sh', 'no sh']
                            output = net_connect.send_config_set(config_commands)
                            # activating shell on SW
                            shell = net_connect.send_command('terminal shell', read_timeout=603)
                            # Checking the logs for confirmation
                            sleep(7)
                            logs = net_connect.send_command('sh log | tail 7', read_timeout=603)
                            if 'Interface ' + neigh['farEndInterface'] + ', ' + 'changed state to down' in logs:
                                print('-> Interface shutdown command executed successfully')
                            elif 'Interface ' + neigh['farEndInterface'] + ', ' + 'changed state to up' in logs:
                                print('-> Interface is UP')

                            if 'no sh' in output:
                                print('interface rebooted')
                        elif restart_int in no_option:
                            print("=> No restart command was sent")

                else:
                    cdp = net_connect.send_command('sh cdp ne ' + neigh['farEndInterface'] + " " + 'de',
                                                   read_timeout=603)
                    neigbor_tl_name = re.search(r'Devi\w+\s\S+\s(.*)', cdp).group(1)
                    if tl_name.lower() == neigbor_tl_name.lower():
                        print('==>*** Input TL and Neighbor TL name match, moving forward ***<==')
                        print(' -> Neighbor name is:', neigbor_tl_name)
                        print(' -> Neighbor name IP is:', re.search(r'IP\sadd\S+\s(.*)', cdp).group(1))
                        platform_nei = re.search(r'Pla\S+\s(\w+\sWS-\S+|\w+)', cdp).group(1)
                        print(' -> Neighbor Platform is:', platform_nei.rstrip(','))
                        local_interface = re.search(r'Inter\w+.\s(\S+)\s+\S+\s\S+\s\S+\s\S+\s(\S+)', cdp).group(1)
                        remote_interface = re.search(r'Inter\w+.\s(\S+)\s+\S+\s\S+\s\S+\s\S+\s(\S+)', cdp).group(2)
                        print(' -> Local Interface is:', local_interface.rstrip(','))
                        print(' -> Remote Interface is:', remote_interface)
                        print('*-----*.*-----*.*-----*.*-----*.*-----*.')
                        # Validating if the TL got a New IP
                        # getting ship CODE
                        new_TL_ip = re.search(r'IP\sadd\S+\s(.*)', cdp).group(1)
                        ship_code = re.search(r'PCL(\w{2})', tl_name).group(1)
                        if cdpData['queryResponse']['entity'][0]['inventoryDetailsDTO']['summary'][
                            'ipAddress'] != new_TL_ip:
                            print(
                                ' -> Tech Locker ' + tl_name + " " + 'is up and reachable, please take into account this device got a NEW IP.'
                                                                     '\n -> Go ahead and ssh into ANY device in ' + ship_code + ',' + ' Instructions here:\n'
                                                                                                                                      '1 -> ssh -l ccl ' + new_TL_ip + '\n'
                                                                                                                                                                       '2 -> sh ver | inc uptime|returned|reload\n'
                                                                                                                                                                       '3 -> Copy and paste the results into the case')
                    else:
                        print("=> Input TL and Neighbor TL did not match, please make sure the TL wasn't replaced")
                        print(' -> Neighbor name is:', neigbor_tl_name)
                        print(' -> Neighbor name IP is:', re.search(r'IP\sadd\S+\s(.*)', cdp).group(1))
                        platform_nei = re.search(r'Pla\S+\s(\w+\sWS-\S+|\w+)', cdp).group(1)
                        print(' -> Neighbor Platform is:', platform_nei.rstrip(','))
                        local_interface = re.search(r'Inter\w+.\s(\S+)\s+\S+\s\S+\s\S+\s\S+\s(\S+)', cdp).group(1)
                        remote_interface = re.search(r'Inter\w+.\s(\S+)\s+\S+\s\S+\s\S+\s\S+\s(\S+)', cdp).group(2)
                        print(' -> Local Interface is:', local_interface.rstrip(','))
                        print(' -> Remote Interface is:', remote_interface)
                        print('*-----*.*-----*.*-----*.*-----*.*-----*.')
                        # Validating if the TL got a New IP
                        # getting ship CODE
                        new_TL_ip = re.search(r'IP\sadd\S+\s(.*)', cdp).group(1)
                        ship_code = re.search(r'PCL(\w{2})', tl_name).group(1)
                        if cdpData['queryResponse']['entity'][0]['inventoryDetailsDTO']['summary'][
                            'ipAddress'] != new_TL_ip:
                            print(
                                ' -> Tech Locker ' + tl_name + " " + 'is up and reachable, please take into account this device got a NEW IP.'
                                                                     '\n -> Go ahead and ssh into ANY device in ' + ship_code + ',' + ' Instructions here:\n'
                                                                                                                                      '1 -> ssh -l ccl ' + new_TL_ip + '\n'
                                                                                                                                                                       '2 -> sh ver | inc uptime|returned|reload\n'
                                                                                                                                                                       '3 -> Copy and paste the results into the case')

                exit(0)

    else:
        print('=> Device ' + tl_name + ' not found it, exiting')
        exit(0)


def send_email(body):
    username = "hsc@mail02.ocean.com"
    password = "123"
    mail_from = "hsc@mail02.ocean.com"
    mail_to = "jose.cristo@hsc.com"
    # mail_subject = "Test Subject"
    mail_subject = 'testing'
    # mail_body = "This is a test message"
    mail_body = body

    mimemsg = MIMEMultipart()
    mimemsg['From'] = mail_from
    mimemsg['To'] = mail_to
    mimemsg['Subject'] = mail_subject
    mimemsg.attach(MIMEText(mail_body, 'plain'))
    connection = smtplib.SMTP(host='mailrelay.qa.ocean.com', port=25)
    ### connection.starttls()
    # connection.login(username, password)
    connection.send_message(mimemsg)
    connection.quit()


def cisco_prime_api_results_devices_IP(device_name, pi_ip):
    requests.packages.urllib3.disable_warnings()

    base_uri = 'https://' + pi_ip + '/webacs/api/v4/'
    user = 'xopsapi'
    password = 'Xops@123'
    # rest_path = '/data/InventoryDetails'
    # rest_path = 'data/AccessPoints?name=eq("Cabana_11")'
    rest_path = 'data/Devices?deviceName=eq("' + device_name + '")'
    url = base_uri + rest_path
    # print(url)

    headers = {'Accept': 'application/json'}
    response = requests.get(url, headers=headers, auth=(user, password), verify=False)
    # print(response.text)
    # print(response.status_code)98745
    acdResp = json.loads(response.text)
    # print(acdResp)

    if acdResp['queryResponse']['@count'] == 1:
        ## print('=> Device Found it, getting details')

        # capturing device ID
        dev_id = acdResp['queryResponse']['entityId'][0]['$']
        # print('id =', dev_id)

        # showing device details
        # device = 'data/Devices/' + dev_id
        device = 'data/Devices/' + dev_id
        url_dev = base_uri + device
        dev_details = requests.get(url_dev, headers=headers, auth=(user, password), verify=False)
        cdpData = json.loads(dev_details.text)
        return cdpData['queryResponse']['entity'][0]['devicesDTO']['ipAddress']


def factory_default_interface(inter, net_connect):
    inter_split = inter.split()
    if '-' in inter:
        print('==> Selecting multiples interface with "-" command')
        config_commands = ['default int range' + " " + inter.lower()]
        result = net_connect.send_config_set(config_commands)
        if 'default int range' + " " + inter.lower() in result:
            print('==> Default factory command executed successfully')
        else:
            print('==> Default factory command not executed')
    #        print(result)

    elif len(inter_split) == 1:
        print("==> Using one interface")
        config_commands0 = ['default int' + " " + inter.lower()]
        #       print(config_commands)
        output0 = net_connect.send_config_set(config_commands0)

        if "set to default" in output0:
            print("Factory default completed")
        else:
            print("not able to factory default the interface")

    elif len(inter_split) > 1:
        print('==> Using not continuous interface range command')
        config_commands1 = ['default int range' + " " + inter.lower()]
        not_contiguous = net_connect.send_config_set(config_commands1)
        if 'default int range' in not_contiguous:
            print('==> Default factory command executed successfully')
        else:
            print('==> Default factory command not executed')
    #        print(not_contiguous)

    else:
        print('Wrong Interface selection')
        exit(0)

    return


def get_temperature(net_connect):
    temperature_dev = net_connect.send_command('show environment temperature | inc OK')
    temperature = net_connect.send_command('show environment temperature | inc Inlet')
    print(temperature_dev)
    print(temperature)


def default_l2_interface(inter, net_connect, vlan_id):
    yes_option = ['yes', 'y']
    no_option = ['no', 'n']

    inter_split = inter.split()
    if '-' in inter:
        print('==> Selecting multiples interface with "-" command')
        config_commands = ['int rang' + " " + inter.lower(), 'switchport mode access', 'spanning-tree portfast',
                           'spanning-tree bpduguard en', 'sw nonegotiate', 'sw acc vlan' + " " + vlan_id]
        output = net_connect.send_config_set(config_commands)
        if 'int rang' + " " + inter.lower() in output:
            print('Commands successfully executed')
            # macros option
            macros = input(
                "==> do you want to disable macros on the interface?, (Y) to continue (N) to cancel:").lower()
            if macros in yes_option:
                disable_macros_range(inter, net_connect)
            elif macros in no_option:
                print("NO changes committed --adios--")
            else:
                print("No changes were made.\nExiting, --BYE--")
        else:
            print('something went wrong, DO NOT CALL JOSE')
    #        print(output)

    elif len(inter_split) == 1:
        print("==> Using one interface")
        config_commands = ['int' + " " + inter.lower(), 'switchport mode access', 'spanning-tree portfast',
                           'spanning-tree bpduguard en', 'sw nonegotiate', 'sw acc vlan' + " " + vlan_id]
        output = net_connect.send_config_set(config_commands)
        if 'sw acc vlan' in output:
            print('==> commands successfully executed\n')
            # macros option
            macros = input(
                "==> do you want to disable macros on the interface?, (Y) to continue (N) to cancel:").lower()
            if macros in yes_option:
                disable_macros(inter, net_connect)
            elif macros in no_option:
                print("NO changes committed --adios--")
            else:
                print("No changes were made.\nExiting, --BYE--")
        else:
            print('something went wrong, DO NOT CALL JOSE')

    elif len(inter_split) > 1:
        print('==> Using not continuous interface range command')
        config_commands = ['int rang' + " " + inter.lower(), 'switchport mode access', 'spanning-tree portfast',
                           'spanning-tree bpduguard en', 'sw nonegotiate', 'sw acc vlan' + " " + vlan_id]
        not_continuous = net_connect.send_config_set(config_commands)
        if 'int rang' + " " + inter.lower() in not_continuous:
            print('==> Commands successfully executed')
            # macros option
            macros = input(
                "==> do you want to disable macros on the interface?, (Y) to continue (N) to cancel:").lower()
            if macros in yes_option:
                disable_macros_range(inter, net_connect)
            elif macros in no_option:
                print("NO changes committed --adios--")
            else:
                print("No changes were made.\nExiting, --BYE--")
        else:
            print('==> something went wrong, DO NOT CALL JOSE')
        #        print(not_continuous)

    else:
        print('==> Wrong Interface selection')
        exit(0)

    return


# show running config only for interfaces loop included
def show_running_config(inter, net_connect):
    inter_split = inter.split()
    #    print(inter)

    if '-' in inter:
        print("Can't show the running config if you use [Gi|Te]x/x/x -X, are you sure you want to continue? ")

    elif 'vlan' in inter:
        print('==> using vlan validator config')
        output = net_connect.send_command('sh run int' + " " + inter)
        print(output)

    elif len(inter_split) >= 1:
        for j in range(len(inter_split)):
            iface = inter_split[j].rstrip(',')
            print("==> Checking current config before change on: " + " " + iface)
            output = net_connect.send_command('sh run int' + " " + iface)
            print(output, '\n')

            print('==> Checking mac-address learned on the ports before change: ', iface)
            output1 = net_connect.send_command('sh mac add int' + " " + iface)
            print(output1, '\n')

            # checking cdp nei
            cdp = net_connect.send_command('sh cdp ne' + " " + iface + " " + 'de')
            if 'Total cdp entries displayed : 0' in cdp:
                print('==> No CDP neighbor on ==>', iface)
            else:
                print('==> Checking CDP neighbor <==')
                nei = get_cdp_neighbor(iface, net_connect)
                print('=> CDP neighbor name is: ', nei[1])
                print('=> CDP neighbor IP is: ', nei[0])
                print('=> CDP neighbor platform is:', nei[2], '\n')

    else:
        pass

    return inter


'''
# only for one interface, general use
def show_running_config_single(inter, net_connect):
    yes_option = ['yes', 'y']
    no_option = ['no', 'n']

    # running command
    output = net_connect.send_command('show run int' + " " + inter)
    print(output)

    factory_default = input("==> do you want to factory default the interface?, (Y) to continue (N) to cancel:").lower()
    if factory_default in yes_option:
        print("==> Sending default int command")
        factory_default_interface(inter, net_connect)

    elif factory_default in no_option:
        print("No factory default interface command applied")
    else:
        print("No factory default interface command applied")

    # Basic L2 interface
    validation = input("==> Do you want to apply the regular L2 config to the interface?, if oK, (Y) to continue (N) "
                       "to cancel:").lower()

    # capturing vlan_id
    vlan_id = input("Whats the VLAN ID?: ")

    # configuring interface
    print('==> Settings interface(s) parameters')

    if validation in yes_option:
        default_l2_interface(inter, net_connect, vlan_id)

    elif validation in no_option:
        print("NO changes committed --adios--")

    else:
        print("No changes were made.\nExiting, --BYE--")
'''


def set_readers_interface(inter, net_connect):
    yes_option = ['yes', 'y']
    no_option = ['no', 'n']
    inter_split = inter.split()

    if '-' in inter:
        print('==> Selecting multiples interface with "-" command')
        config_commands = ['int' + " " + inter, 'desc [RDR] MED-READER', 'spanning-tree portfast',
                           'spanning-tree bpduguard en', 'switchport nonegotiate', 'sw mod acc', 'sw acc vlan 1310',
                           'power inline static max 30000']
        output = net_connect.send_config_set(config_commands)
        if 'int rang' + " " + inter.lower() in output:
            print('Commands successfully executed')
            # macros option
            macros = input(
                "==> do you want to disable macros on the interface?, (Y) to continue (N) to cancel:").lower()
            if macros in yes_option:
                disable_macros(inter, net_connect)
            elif macros in no_option:
                print("NO changes committed --adios--")
            else:
                print("No changes were made.\nExiting, --BYE--")
        else:
            print('something went wrong, DO NOT CALL JOSE')
    #        print(output)

    elif len(inter_split) == 1:
        print("==> Using one interface")
        config_commands = ['int' + " " + inter, 'desc [RDR] MED-READER', 'spanning-tree portfast',
                           'spanning-tree bpduguard en', 'sw mod acc', 'sw acc vlan 1310',
                           'power inline static max 30000']
        output = net_connect.send_config_set(config_commands, read_timeout=603)
        if 'sw acc vlan' in output:
            print('Commands successfully executed')
            # macros option
            macros = input(
                "==> do you want to disable macros on the interface?, (Y) to continue (N) to cancel:").lower()
            if macros in yes_option:
                disable_macros(inter, net_connect)
            elif macros in no_option:
                print("NO changes committed --adios--")
            else:
                print("No changes were made.\nExiting, --BYE--")
        else:
            print('something went wrong, DO NOT CALL JOSE')
    #        print(output)

    elif len(inter_split) > 1:
        print('==> Using not continuous interface range command')
        config_commands = ['int' + " " + inter, 'desc [RDR] MED-READER', 'spanning-tree portfast',
                           'spanning-tree bpduguard en', 'sw mod acc', 'sw acc vlan 1310',
                           'power inline static max 30000']
        not_continuous = net_connect.send_config_set(config_commands, read_timeout=603)
        if 'int rang' + " " + inter.lower() in not_continuous:
            print('==> Commands successfully executed')
            # macros option
            macros = input(
                "==> do you want to disable macros on the interface?, (Y) to continue (N) to cancel:").lower()
            if macros in yes_option:
                disable_macros_range(inter, net_connect)
            elif macros in no_option:
                print("NO changes committed --adios--")
            else:
                print("No changes were made.\nExiting, --BYE--")
        else:
            print('==> something went wrong, DO NOT CALL JOSE')
    #        print(not_continuous)

    else:
        print('==> Wrong Interface selection')
        exit(0)

    return


def disable_macros(inter, net_connect):
    print(' Disabling macros on interface(s)')
    config_commands = ['int' + " " + inter, 'no macro auto proc']
    output = net_connect.send_config_set(config_commands)
    if 'no macro auto' in output:
        print('==> macros disabled successfully\n')
    else:
        print('==> no macros command was sent\n')


def disable_macros_range(inter, net_connect):
    print(' Disabling macros on interface(s)')
    config_commands = ['int range' + " " + inter, 'no macro auto proc']
    output = net_connect.send_config_set(config_commands)
    if 'no macro auto' in output:
        print('==> macros disabled successfully\n')
    else:
        print('==> no macros command was sent\n')


def bounce_interface(inter, net_connect):
    inter_split = inter.split()
    if '-' in inter:
        print('==> Selecting multiples interface with "-" command')
        config_commands = ['int rang' + " " + inter.lower(), 'sh', 'no sh']
        output = net_connect.send_config_set(config_commands)
        if 'no sh' in output:
            print('interface rebooted')
        else:
            print('something went wrong, DO NOT CALL JOSE')
    #        print(output)

    elif len(inter_split) == 1:
        print("==> Using one interface")
        config_commands = ['int rang' + " " + inter.lower(), 'sh', 'no sh']
        output = net_connect.send_config_set(config_commands)
        if 'no sh' in output:
            print('interface rebooted')
        else:
            print('something went wrong, DO NOT CALL JOSE')
    #        print(output)

    elif len(inter_split) > 1:
        print('==> Using not continuous interface range command')
        config_commands = ['int rang' + " " + inter.lower(), 'sh', 'no sh']
        not_continuous = net_connect.send_config_set(config_commands)
        #        print(not_continuous)
        if 'no sh' in not_continuous:
            print('interface rebooted')
        else:
            print('==> something went wrong, DO NOT CALL JOSE')
    #        print(not_continuous)

    else:
        print('==> Wrong Interface selection')
        exit(0)

    return


def show_results(inter, net_connect):
    inter_split = inter.split()
    if len(inter_split) >= 1:
        for j in range(len(inter_split)):
            iface = inter_split[j].rstrip(',')
            print("==> Results after config")
            output = net_connect.send_command('sh run int' + " " + iface)
            print(output, '\n')
    else:
        pass

    return


def check_arp_from_ip_or_mac(mac, net_connect, vlan_id):
    print("==> Searching on the ARP table")
    output = net_connect.send_command('sh ip arp vlan' + " " + vlan_id + " " + "| inc" + " " + mac)
    #    print(output, '\n')

    ip = re.search(r'(\d{1,3})\.(\d{1,3})\.(\d{1,3})\.(\d{1,3})', output)
    l2 = re.search(r'([a-f]|[0-9]){4}\.([a-f]|[0-9]){4}\.([a-f]|[0-9]){4}', output)

    if ip.group() in output:
        print('The IP address associated with the mac is: ' + " " + ip.group())
        print('The mac-add associated with the IP is: ' + " " + l2.group())
    else:
        print('mac-add or IP address not found it')

    return


def save_config(net_connect):
    print("==> saving config")
    output = net_connect.send_command('wr mem', read_timeout=603)
    print(output)


def check_mac_add_interface_dst(mac, net_connect):
    print("==> Searching mac-address")
    output = net_connect.send_command('sh mac add ad' + " " + mac)

    #    print("==> grabbing the mac-add")
    result = re.search(r'((Gi)\d/.*|(Po\d+)|(Te\d+))', output)  # re.search is just for the first entry
    # result = re.finditer(r'10\.[0-2]{1,3}\S+', output)

    if 'Po' in result.group():

        print('==> The requested mac is behind the following interface: ', result.group(), '==> checking who is '
                                                                                           'behind\n')

        # grab Po number
        #        print('this is result', result.group())
        po = result.group()
        #        print('this is po', po)
        # checking port-channel
        po_check = net_connect.send_command('sh etherch summ | inc' + " " + po)
        # grabbing interface from ether-channel summary output
        po_result = re.search(r'((Gi)\d/\d/\d+|(Te\d/\d/\d+))', po_check)
        po_result_str = po_result.group()
        #        print('this is po_result', po_result_str)
        cdp = net_connect.send_command('sh cdp ne' + " " + po_result_str + " " + 'de')
        cdp_name = re.search(r'Dev.*', cdp)
        cdp_name1 = re.search(r'[^Device\sID:].*', cdp_name.group())
        cdp_ip = re.search(r'IP.*', cdp)
        cdp_ip1 = re.search(r'\d.*', cdp_ip.group())

        print('The mac-add is behind the following switch: ' + " " + cdp_name1.group())
        print('The IP for the neighbor is: ' + " " + cdp_ip1.group(), '\n')
        if not cdp_name1:
            print('==> No more neighbors')
        else:
            print('Connecting to neighbor ==>', cdp_name1.group())
            # trying to connect to neighbor
            '''
            cdp_ip_neigh = cdp_ip1.group()
            JC = credentials_reconnect(cdp_ip_neigh)
            net_connect = ConnectHandler(**JC)
            net_connect.enable()

            # getting info
            reconnect_and_check_mac_add_interface_dst(mac, net_connect)
            '''
            cdp_ip_neigh = cdp_ip1.group()
            JC = credentials_reconnect(cdp_ip_neigh)
            net_connect = ConnectHandler(**JC)
            net_connect.enable()

            # getting info
            check_mac_add_interface_dst(mac, net_connect)

    elif 'Gi' in result.group():

        # checking mac-add learned
        output = net_connect.send_command('sh mac add int' + " " + result.group())
        cdp = net_connect.send_command('sh cdp ne' + " " + result.group() + " " + 'de')
        # filtering total mac number
        mac_total = re.search(r'\s\d+$.*', output)
        mac_total = int(mac_total.group().strip())
        #        print('total macs are', mac_total)
        #        print('this is the interface', result)
        if 'Total cdp entries displayed : 0' in cdp:
            print('==> No more cdp neighbors', '==> The provided mac-add is behind: ', result.group())
            exit(0)
        else:
            nei = get_cdp_neighbor(result.group(), net_connect)
            print('The mac-add is behind the following switch: ' + " " + nei[1])
            print('The IP for the neighbor is: ' + " " + nei[0], '\n')
            print('Connecting to neighbor ==>', nei[1])

            JC = credentials_reconnect(nei[0])
            net_connect = ConnectHandler(**JC)
            net_connect.enable()

        # getting info
        check_mac_add_interface_dst(mac, net_connect)

    else:
        exit(0)

    return


# getting cdp_nei_IP
def get_cdp_neighbor(inter, net_connect):
    # getting cdp neighbor
    cdp = net_connect.send_command('sh cdp ne' + " " + inter + " " + 'de', read_timeout=903     )

    if 'Total cdp entries displayed : 0' in cdp:
        print('==> No CDP neighbor')
        pass
    else:
        cdp = net_connect.send_command('sh cdp ne' + " " + inter + " " + 'de')
        cdp_name = re.search(r'Dev.*', cdp)
        cdp_name1 = re.search(r'[^Device\sID:].*', cdp_name.group())
        # validating if IPv4 field exist
        if re.search(r'IP\s\w+.*', cdp):
            cdp_ip = re.search(r'IP\s\w+.*', cdp)
            cdp_ip1 = re.search(r'\d.*', cdp_ip.group())
        else:
            var = 'none'
            cdp_ip1 = re.search(r'\w+', var)

        # added cisco and linux for Platform
        platform = re.search(r'Plat\w+.\s+\w+\s\w+.\w+|Plat\w+.\s\w+', cdp)
        # platform = re.search(r'Plat\w+.\s+\w+\s\w+.\w+', cdp) # only for cisco
        #  added cisco or linux
        platform1 = re.search(r'(ci.*)|(Li\w+)', platform.group())
        # platform1 = re.search(r'(ci.*)', platform.group()) # only for cisco

        return cdp_ip1.group(), cdp_name1.group(), platform1.group()


def get_lldp_neighbor(inter, net_connect):
    # getting lldp info
    lldp = net_connect.send_command('sh cdp ne' + " " + inter + " " + 'de')

    if 'Total entries displayed: 0' in lldp:
        print('==> No LLDP neighbor <==')
        pass
    else:
        lldp_neighbor = net_connect.send_command('sh lldp ne' + " " + inter + " " + 'de')
        # print(lldp_neighbor)
        lldp_name = re.search(r'Chas.*', lldp_neighbor)
        lldp_name1 = re.search(r'[^Chassis\sid:].*', lldp_name.group())
        sys_name = re.search(r'Sys\w+\sN\w+.*', lldp_neighbor)
        sys_name1 = re.search(r'[^System Name:].*', sys_name.group())
        description = re.search(r'System Description:\s[\r\n]+([^\r\n]+)', lldp_neighbor)
        # print(description.group())
        # validating if IPv4 field exist
        if re.search(r'IP:\s\w+.*', lldp_neighbor):
            lldp_ip = re.search(r'IP:\s.*', lldp_neighbor)
            lldp_ip1 = re.search(r'\d.*', lldp_ip.group())
        else:
            var = 'none'
            lldp_ip1 = re.search(r'\w+', var)

        return lldp_name1.group(), lldp_ip1.group(), sys_name1.group(), description.group()


# function to reconnect to a device!!!
def credentials_reconnect(cdp_ip1):
    if '10.5.144' in cdp_ip1 or '10.126.70' in cdp_ip1 or '10.5.160' in cdp_ip1 or '10.126.78.130' in cdp_ip1:
        print('=> Using CCL Credentials')
        JC = {
            'device_type': 'cisco_ios',
            'ip': cdp_ip1,
            'username': 'ccl',
            'password': 'N@v!gaT!nG~',
            'timeout': 29,
            'global_delay_factor': 7,
        }
        return JC
    else:
        print('=> Using TACACS Credentials')
        JC = {
            'device_type': 'cisco_ios',
            'ip': cdp_ip1,
            'username': 'jcr8398',
            'password': 'VRRP cce2010',
            'timeout': 29,
            'global_delay_factor': 6,

        }
        return JC


# show if interface is up or down
def check_interface_status(inter, net_connect):
    print('==> Checking interface status( UP || DOWN )')

    if 'gi' in inter or 'te' in inter or 'Te' in inter or 'Po' in inter or 'vl' in inter:
        print('=> Using IOS interface')
        output = net_connect.send_command('sh int' + " " + inter + " " + '| inc line')
        print(output, '\n')

    elif 'eth' in inter:
        print('=> Using Nexus interface')
        output = net_connect.send_command('sh int' + " " + inter + " " + '| inc Link')
        print(output, '\n')

    else:
        pass

    return


def get_hostname_only(net_connect):
    name = net_connect.send_command('sh run | inc hostname')
    hostname = re.search(r'host\w+\s(.*)', name)
    # print(hostname.group(1))
    return hostname.group(1)


def get_stackwise_size(net_connect):
    stack = net_connect.send_command('sh switch')
    stackwise = re.findall(r'(\d)\s+\bActive\b|(\d)\s+\bStandby\b|(\d)\s+\bMember\b', stack)
    print(len(stackwise))


def get_hostname(net_connect):
    output = net_connect.send_command('sh ver | inc Nexus')
    if output:
        print('==> Nexus Device Detected <==')
        get_device_date_time(net_connect)
        nexus = net_connect.send_command('sh system uptime')
        # getting hostname
        nexus_hostname = net_connect.send_command('sh run | inc hostname')
        hostname = re.search(r'host\w+\s(.*)', nexus_hostname)
        print(nexus, 'Hostname =>', hostname.group(1))

    else:
        print('==> IOS Device Detected <==')
        get_device_date_time(net_connect)
        output = net_connect.send_command('sh ver | inc uptime|Uptime|Last')
        print(output)


def check_interface_details_and_po(inter, net_connect):
    # checking running int config to validate if INT is member of a PO
    if 'po' in inter or 'port-channel' in inter:
        print('=> PO interface selected')
        po_int = net_connect.send_command(
            'sh int' + " " + inter + " " + '| inc Desc|Hard|MTU|line|media|Input|CRC|port-channel')
        print(po_int)

    elif 'gi' in inter or 'te' in inter or 'vl' in inter:
        print('=> IOS device detected')
        # getting device name
        dev_name = net_connect.send_command('sh run' + " " + '| inc hostname')
        host = re.search(r'host\w+\s(.*)', dev_name)
        print('=> Hostname =>', host.group(1))
        # print('==> Getting interface details')
        interface = net_connect.send_command('sh int' + " " + inter + " " + '| inc Desc|Hard|MTU|line|media|Input|CRC')
        print(interface)
        print('*---*-*---*-*---*-*---*-')
        run_int = net_connect.send_command('sh run int' + " " + inter)
        # checking running int config to validate if INT is member of a PO
        if 'channel-group' in run_int:
            po_int = re.search(r'cha\w+-gr\w+\s(\d+)', run_int)
            print('=> Interface' + " " + inter + " " + 'belong to PO' + " " + po_int.group(1))
            po_int_info = net_connect.send_command(
                'sh int' + " " + 'po' + po_int.group(1) + " " + '| inc Desc|Hard|MTU|line|media|Input|CRC')
            print(po_int_info)
        else:
            print('=> Interface does not belong to a PO')

    elif 'eth' in inter:
        print('=> NXOS device detected')
        nxos_name = net_connect.send_command('sh run' + " " + '| inc hostname')
        nxos_host = re.search(r'host\w+\s(.*)', nxos_name)
        print('=> Hostname =>', nxos_host.group(1))
        nxos = net_connect.send_command(
            'sh int' + " " + inter + " " + '| inc Desc|Hard|MTU|Etherne|media|Port|resets|flapped|Belongs')
        print(nxos)
        print('*---*-*---*-*---*-*---*-')
        # checking running int config to validate if INT is member of a PO
        if 'Belongs' in nxos:
            nx_po = re.search(r'Belo\w+\s\w+\s(\w+)', nxos)
            print('=> Interface' + " " + inter + " " + 'belong to' + " " + nx_po.group(1))
            nxos_details = net_connect.send_command(
                'sh int' + " " + nx_po.group(1) + " " + '| inc Desc|Hard|MTU|vPC|media|Port|port-channel')
            print(nxos_details)
    else:
        print('=> NO INT HIT')


def get_credentials_and_interface():
    #    inter = 0
    #    if isInterfaceNeeded:
    #        inter = str(input("==> Enter the interface(s), more than one use ,space (Gi1/0/1, Gi3/0/33) <==: "))
    print('==> Using hardcoded JC credentials <==')
    IP = input("Give me the device IP: ")
    # USERNAME = input("What's the username: ")
    # PASS = getpass.getpass()

    if '10.5.144' in IP or '10.126.78.130' in IP or '10.5.160' in IP or '10.126.140.125' in IP:
        print('=> Using CCL Credentials')
        JC = {
            'device_type': 'cisco_ios',
            'ip': IP,
            'username': 'ccl',
            'password': 'N@v!gaT!nG~',
            'timeout': 41,
            'global_delay_factor': 7,
            'banner_timeout': 303,
            'conn_timeout': 650,
            'read_timeout_override': 701,
        }
        return JC
    else:
        print('=> Using TACACS Credentials')
        JC = {
            'device_type': 'cisco_ios',
            'ip': IP,
            'username': 'jcr8398',
            'password': 'VRRP cce2010',
            'timeout': 41,
            'global_delay_factor': 9,
            'banner_timeout': 77,
            'conn_timeout': 650,

        }
        return JC

    '''
    JC = {
        'device_type': 'cisco_ios',
        'ip': IP,
        'username': 'jcr8398',
        'password': 'STP cce2010',
        
    }
    return JC
    '''


def if_credential_connection(ip):
    #    inter = 0
    #    if isInterfaceNeeded:
    #        inter = str(input("==> Enter the interface(s), more than one use ,space (Gi1/0/1, Gi3/0/33) <==: "))

    JC = {
        'device_type': 'cisco_ios',
        'ip': ip,
        'username': 'jcr8398',
        'password': 'VRRP cce2010',
        'timeout': 29,
        'global_delay_factor': 6,
        'banner_timeout': 77,
        'conn_timeout': 650,
    }
    return JC


def panos_credentials():
    IP = input("Give me the PAN-OS device IP: ")
    # USERNAME = input("What's the username: ")
    # PASS = getpass.getpass()

    panos = {
        'device_type': 'paloalto_panos',
        'ip': IP,
        'username': 'Read_Only',
        'password': 'S$@!L!nG!12',
        # 'password': 'R0-Only1',
        'timeout': 29,
        'global_delay_factor': 7,
    }
    return panos


def get_potential_for_disaster(isIP, net_connect):
    # getting CDP neighbor
    cdp = net_connect.send_command('sh cdp ne', read_timeout=603)
    # filtering TL
    tl_s = re.findall(r'WS-C3560C', cdp)
    # Filtering Phones
    phone = re.findall(r'Mitel', cdp)
    # Filtering APs
    ap_s = re.findall(r'C9120AXI|C9130AXI|AP380', cdp)
    # Filtering Readers
    readers = re.findall(r'Linux', cdp)
    # Printing
    print('*...*.*...*.*...*.*...*')
    # Getting the time
    print('==> Local time & Date =', get_time_date()[0], '=> Time =', get_time_date()[1])
    # Getting device name
    print('=> Hostname:', get_hostname_only(net_connect), '=>', isIP.rstrip())
    # Printing the Model and OS
    print('=> Model:', get_ios_nxos_version_model(net_connect)[0])
    print('=> OS or Code:', get_ios_nxos_version_model(net_connect)[1])
    # checking the sack size
    stack = net_connect.send_command('sh switch')
    stack_size = re.findall(r'.*Ready', stack)
    if len(stack_size) == 1:
        print('=> There is', len(stack_size), 'switch on the stack')
    elif len(stack_size) > 1:
        print('=> There are', len(stack_size), 'switches on the stack')
    # print total of cabin affected
    print('=> Total of potentially Cabins affected', len(tl_s) * 2)
    # print total of phones
    print('=> Total of potentially Phones affected', len(phone))
    # print total APs
    print('=> Total of potentially APs affected', len(ap_s))
    # print total readers
    print('=> Total of potentially Readers affected', len(readers))


def get_idf_switches_from_fzs(net_connect):
    # Getting hostname
    name = get_hostname_only(net_connect)
    name_detail = re.search(r'^\w{3}(\w{2})', name).group(1)
    # getting interfaces UP
    int_connected = net_connect.send_command('sh cdp ne', read_timeout=707)
    # filtering the output to grab interfaces
    interfaces_up = re.findall(r'PCL\S+\s+(\S+\s\S+)', int_connected)
    # Excel info
    v_boldFont = openpyxl.styles.Font(bold=True)
    v_centerAlignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrapText=True)
    wb = load_workbook(name_detail + '_IDF_Inventory.xlsx')
    # wb = Workbook()
    ws = wb.active
    # wb.create_sheet(get_ship_name)
    # ws.append(['Neigh_Name', 'Neigh_IP', 'Neigh_Platform', 'Local_Interface', 'Neigh_Interface'])
    ws.title = 'IDF_Inventory'  # creating sheet title

    for c in interfaces_up:
        # checking cdp neighbor
        print('=> Checking cdp neighbor in Interface ' + str(c))
        cdp = net_connect.send_command('sh cdp ne ' + str(c) + ' de', read_timeout=707)
        cdp_name = re.search(r'Dev\w+\s\S+\s+([^.]+)', cdp).group(1)
        if cdp_name and 'IDF' in cdp_name or 'ID' in cdp_name:
            print('=>yes,IDF')
            cdp_name = re.search(r'Dev\w+\s\S+\s+([^.]+)', cdp).group(1)
            cdp_ip = re.search(r'Ent\S+\s\S+\s[\r\n]+(\s+\S+\s+\S+\s(.*))', cdp).group(2)
            cdp_platform = re.search(r'Plat\S+\s(\S+\s\S+|\S+)', cdp).group(1).rstrip(',')
            cdp_interface = re.search(r'Int\S+\s(\S+)', cdp).group(1).rstrip(',')
            cdp_remote_interface = re.search(r'Por\w+\sID\s\S+\s\S+\s(\S+)', cdp).group(1)
            # filling out the Exel
            ws.append([cdp_name, cdp_ip, cdp_platform, cdp_interface, cdp_remote_interface])

        else:
            print('not an IDF')

    # Bold and align to the Center
    for cell in ws[1]:
        cell.font = v_boldFont
        cell.alignment = v_centerAlignment

    # saving the Excel
    wb.save(name_detail + '_IDF_Inventory.xlsx')


def get_ios_nxos_version_model(net_connect):
    show_ver = net_connect.send_command('sh ver', read_timeout=603)
    if 'NXOS' in show_ver:
        nxos_version = re.search(r'NXOS.\sver\S+\s(.*)', show_ver)
        nxos_model = re.search(r'Hardware[\r\n]\s+\S+\s(\S+\s\S+)', show_ver)
        return nxos_model.group(1), nxos_version.group(1)

    elif 'IOS' in show_ver:
        model = re.search(r'SW\sIma\S+\s+Mo\w+\s+[\r\n]+([^\r\n]+)[\r\n]\S+\s+\S\s\S+\s+(\S+)\s+(\S+)', show_ver).group(
            2)
        os_ver = re.search(r'SW\sIma\S+\s+Mo\w+\s+[\r\n]+([^\r\n]+)[\r\n]\S+\s+\S\s\S+\s+(\S+)\s+(\S+)',
                           show_ver).group(3)
        return model, os_ver


def get_ios_nxos_name(net_connect):
    version = net_connect.send_command('sh run | inc hostn')
    dev_name = re.search(r'hos\w+\s(.*)', version).group(1)
    print('=> Device Name:', dev_name)


def check_bgp_network(bgp_route, net_connect):
    print('==> Checking bgp network||IP on BGP neighbor table\n')
    # getting the info
    output = net_connect.send_command('sh ip bgp all summa | inc' + " " + bgp_route)

    # filtering if more than one result.
    output_filter = re.findall(r'(\d+\.\d+\.\d+\.\d+\s+\d\s+\d+.\d+.*)', output)

    if len(output_filter) >= 2:
        print('==> More than one neighbor <==')
        for j in range(len(output_filter)):
            # tired of the..FUCK...G errors, searching for Active or something
            not_up = re.search(r'\((.+)\)|\b(\w+)+\b$', output_filter[j])
            # print(not_up.group())
            # if ('Idle' or 'never' or 'Active') in output_filter[j]:
            if 'Idle' in not_up.group() or 'never' in not_up.group() or 'Active' in not_up.group() or 'Admin' in not_up.group():
                if 'Admin':
                    print('=> The neighbor ' + " " + bgp_route + " " + 'is Administratively shut down <==\n')
                else:
                    bgp_ip = re.search(r'\d[1-9]{1,3}\.\d+\.\d+\.\d+', output_filter[j])
                    print('=> The neighbor' + " " + bgp_route + " " + 'is down <==')
                    # checking if the provided network belong to VRF guest_internet
                    guest_internet_vrf = net_connect.send_command("sh run | sec Guest_Internet")
                    if bgp_route in guest_internet_vrf:
                        print(
                            '=> The provided neighbor ' + bgp_route + " Belongs to guest_internet VRF and isn't in use anymore")

            else:
                # testing with vrf details #
                vrf_list = ['Trident', 'Trident-SDN', 'MDL-PAX', 'Ocean', 'Voice', 'MDL-CREW']
                global_routing = net_connect.send_command('sh ip bgp neighbor' + " " + bgp_route)
                if 'No such neighbor' in global_routing:
                    for c in range(len(vrf_list)):
                        vrf = net_connect.send_command(
                            'sh ip bgp vpnv4 vrf' + " " + str(vrf_list[c]) + " " + 'neighbors' + " " + bgp_route)
                        # print('vrf', vrf)
                        if 'No such neighbor' in vrf:
                            pass
                        else:
                            # getting bgp version
                            bgp_nei = re.search(r'BGP\sne.*', vrf)
                            bgp_ver = re.search(r'BGP\sv.*', vrf)
                            bgp_state = re.search(r'BGP\ss.*', vrf)
                            print('=> ' + bgp_nei.group())
                            print('=> ' + bgp_ver.group())
                            print('=> ' + bgp_state.group())
                ## end Testing ##

                print('==> more than one BGP neighbor received <==')
                print('=> Date =', get_time_date()[0], '=> Time =', get_time_date()[1])
                get_ios_nxos_name(net_connect)
                result_first = re.search(r'(\d+\.\d+\.\d+\.\d+\s+\d\s+\d+.\d+)', output)
                result_2nd = re.search(r'(\d{2}:\d{2}:\d+)|(\d{1,2}[a-z]\w+)', output)
                result_3rd = re.search(r'\d+$', output)
                print('Neighbor        V        AS   Up/Down PfxRcd')
                print(result_first.group() + " " + result_2nd.group(), result_3rd.group(), '\n')

    else:
        print('==> Only one neighbor received <==')
        print('=> Date =', get_time_date()[0], '=> Time =', get_time_date()[1])
        get_ios_nxos_name(net_connect)
        # testing with vrf details #
        vrf_list = ['Trident', 'Trident-SDN', 'MDL-PAX', 'Ocean', 'Voice', 'MDL-CREW']
        global_routing = net_connect.send_command('sh ip bgp neighbor' + " " + bgp_route)
        if 'No such neighbor' in global_routing:
            for c in range(len(vrf_list)):
                vrf = net_connect.send_command(
                    'sh ip bgp vpnv4 vrf' + " " + str(vrf_list[c]) + " " + 'neighbors' + " " + bgp_route)
                # print('vrf', vrf)
                if 'No such neighbor' in vrf or 'Unknown VRF' in vrf:
                    pass
                else:
                    # getting bgp version
                    bgp_nei = re.search(r'BGP\sne.*', vrf)
                    bgp_ver = re.search(r'BGP\sv.*', vrf)
                    bgp_state = re.search(r'BGP\ss.*', vrf)
                    print('=> ' + bgp_nei.group())
                    print('=> ' + bgp_ver.group())
                    print('=> ' + bgp_state.group())

        ## end Testing ##

        for j in range(len(output_filter)):
            # tired of the errors..FU......K, searching for Active or something
            not_up1 = re.search(r'\((.+)\)|\b(\w+)+\b$', output_filter[j])
            # print('not_up', not_up1.group())
            # print('=>', not_up1.group())

            if 'Active' in not_up1.group() or 'never' in not_up1.group() or 'Idle' in not_up1.group() or 'Admin' in not_up1.group():
                bgp_ip = re.search(r'\d[1-9]{1,3}\.\d+\.\d+\.\d+', output_filter[j])
                print('=> The neighbor' + " " + bgp_route + " " + 'is down <==')
                # checking if the provided network belong to VRF guest_internet
                guest_internet_vrf = net_connect.send_command("sh run | sec Guest_Internet")
                if bgp_route in guest_internet_vrf:
                    print(
                        '=> The provided neighbor ' + bgp_route + " Belongs to guest_internet VRF and isn't in use anymore")

            else:
                result_first = re.search(r'(\d+\.\d+\.\d+\.\d+\s+\d\s+\d+.\d+)', output)
                result_2nd = re.search(r'(\d{2}:\d{2}:\d+)|(\d{1,2}[a-z]\w+)', output)
                result_3rd = re.search(r'\d+$', output)
                print('Neighbor        V        AS   Up/Down PfxRcd')
                print(result_first.group() + " " + result_2nd.group(), result_3rd.group())


def check_ap_cdp_neighbor(ap_name, net_connect):
    print("==> Getting CDP neighbor for: ", ap_name)
    output = net_connect.send_command("show ap cdp neighbors ap-name" + " " + ap_name)
    print('\n')

    if output:
        # filtering CDP switch name
        cdp = re.search(r'(PCL\w+.\d{3}-\w+)|(PCL\w+.\w+)', output)
        # getting nei IP add
        nei_ip_add = re.search(r'IP\sadd.*', output)
        nei_ip = re.search(r'\d.*', nei_ip_add.group())

        # Filtering nei interface
        interface = re.search(r'(Gi.*)|(Te.*)', output)

        # printing results:
        print('The AP cdp neighbor is: ', cdp.group())
        print('The neighbor IP is: ', nei_ip.group())
        print('The neighbor interface is: ', interface.group())

    else:
        print("I couldn't find any neighbor")

    return


def check_bgp_neighbors_all(net_connect):
    print('==> Getting all the Neighbors: <== \n')

    bgp_all_str = ''

    # showing neighbors
    output = net_connect.send_command('sh ip bgp all summa')
    if not output:
        print('No BGP Neighbors')
    else:
        # bgp_all = re.search(r'\d+\.\d+\.\d+\.\d+\s+4.*', output)
        bgp_all = re.findall(r'\d+\.\d+\.\d+\.\d+\s+4.*', output)
        # bgp_ip = re.findall(r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d+', output)
        # bgp_prefix = re.findall(r'\w+$', bgp_all_str)

        for j in range(len(bgp_all)):
            bgp_all_str += bgp_all[j] + '\n'

    print('Neighbor  V      AS MsgRcvd MsgSent   TblVer  InQ OutQ Up/Down  State/PfxRcd')
    print(bgp_all_str)

    return


def check_bgp_ocean_table(net_connect):
    # validating if the nei is on VRF (XiC and few others exceptions)
    tunnel_int = net_connect.send_command('sh run int Tunnel20124 | inc forwarding')

    if 'vrf' in tunnel_int:
        vrf = re.search(r'\b((\w+)|(\w+-\w+))+\b$', tunnel_int)
        print('==> Using neighbor inside VRF : ', vrf.group())
        output = net_connect.send_command('sh ip bgp vpnv4 vrf Trident-SDN')
        print(output)
    else:
        print('==> Using non VRF interface')
        non_vrf = net_connect.send_command("sh ip bgp")
        print(non_vrf)

    return


def check_bgp_ocean_network_specific(bgp_net, net_connect):
    # validating if the nei is on VRF (XiC and few other's exception)
    tunnel_int = net_connect.send_command('sh run int Tunnel20124 | inc forwarding')

    if 'vrf' in tunnel_int:
        vrf = re.search(r'\b((\w+)|(\w+-\w+))+\b$', tunnel_int)
        print('==> Using neighbor inside VRF : ', vrf.group())
        output = net_connect.send_command('sh ip bgp vpnv4 vrf Trident-SDN' + " " + bgp_net)
        print(output)
    else:
        print('==> Using non VRF interface')
        non_vrf = net_connect.send_command("sh ip bgp" + " " + bgp_net)
        print(non_vrf)


def check_bgp_ocean_received_routes(aws_ip, net_connect):
    # validating if the nei is on VRF (XiC and few other's exception)
    tunnel_int = net_connect.send_command('sh run int Tunnel20124 | inc forwarding')

    if 'vrf' in tunnel_int:
        vrf = re.search(r'\b((\w+)|(\w+-\w+))+\b$', tunnel_int)
        print('==> Using neighbor inside VRF : ', vrf.group())
        output = net_connect.send_command('sh ip bgp vpnv4 vrf Trident-SDN neighbors' + " " + aws_ip + " " + 'received'
                                                                                                             '-routes')
        print(output)
    else:
        print('==> Using non VRF interface')
        non_vrf = net_connect.send_command("sh ip bgp neighbors" + " " + aws_ip + " " + 'received-routes')
        print(non_vrf)

    return


def check_bgp_ocean_sent_routes(aws_ip, net_connect):
    # validating if the nei is on VRF (XiC and few other's exception)
    tunnel_int = net_connect.send_command('sh run int Tunnel20124 | inc forwarding')

    if 'vrf' in tunnel_int:
        vrf = re.search(r'\b((\w+)|(\w+-\w+))+\b$', tunnel_int)
        print('==> Using neighbor inside VRF : ', vrf.group())
        output = net_connect.send_command(
            'sh ip bgp vpnv4 vrf Trident-SDN neighbors' + " " + aws_ip + " " + 'advertised-routes')
        print(output)
    else:
        print('==> Using non VRF interface')
        non_vrf = net_connect.send_command("sh ip bgp neighbors" + " " + aws_ip + " " + 'advertised-routes')
        print(non_vrf)

    return


def check_bgp_neighbor_ocean_details(aws_ip, net_connect):
    # validating if the nei is on VRF (XiC and few other's exception)
    tunnel_int = net_connect.send_command('sh run int Tunnel20124 | inc forwarding')
    if 'vrf' in tunnel_int:
        vrf = re.search(r'\b((\w+)|(\w+-\w+))+\b$', tunnel_int)
        print('==> Using neighbor inside VRF : ', vrf.group())
        nei_details = net_connect.send_command("sh ip bgp vpnv4 vrf Trident-SDN neighbors" + " " + aws_ip)
        # filtering results
        bgp_nei = re.search(r'^BGP.*', nei_details)
        inherits = re.search(r'Inheri.*', nei_details)
        bgp_ver = re.search(r'BGP\svers.*', nei_details)
        bgp_state = re.search(r'BGP\ssta.*', nei_details)
        prefix_current = re.search(r'Prefixes\sCurr\w+.\s+\d+\s+\d+', nei_details)
        rcvd_prefixes = re.search(r'\b(\d*)\b$', prefix_current.group())
        sent_prefixes_total = re.search(r'Prefixes\sCurr\w+.\s+\d*', prefix_current.group())
        sent_prefixes = re.search(r'\b(\d*)\b$', sent_prefixes_total.group())
        print('#-#-#-#-#-#-#-#-#-#_#-#')
        # printing results
        print('Total prefixes received = ', rcvd_prefixes.group())
        print('Total prefixes sent = ', sent_prefixes.group())
        print('#-#-#-#-#-#-#-#-#-#_#-#')
        print(bgp_nei.group())
        print(inherits.group())
        print(bgp_ver.group())
        print(bgp_state.group())

    else:
        print('==> Using non VRF interface')
        nei_details = net_connect.send_command("sh ip bgp neighbors" + " " + aws_ip)
        # filtering results
        bgp_nei1 = re.search(r'^BGP.*', nei_details)
        # inherits1 = re.search(r'Inheri.*', nei_details)
        bgp_ver1 = re.search(r'BGP\svers.*', nei_details)
        bgp_state1 = re.search(r'BGP\ssta.*', nei_details)
        prefix_current1 = re.search(r'Prefixes\sCurr\w+.\s+\d+\s+\d+', nei_details)
        rcvd_prefixes1 = re.search(r'\b(\d*)\b$', prefix_current1.group())
        sent_prefixes_total1 = re.search(r'Prefixes\sCurr\w+.\s+\d*', prefix_current1.group())
        sent_prefixes1 = re.search(r'\b(\d*)\b$', sent_prefixes_total1.group())
        print('#-#-#-#-#-#-#-#-#-#_#-#')
        # printing results
        print('Total prefixes received = ', rcvd_prefixes1.group())
        print('Total prefixes sent = ', sent_prefixes1.group())
        print('#-#-#-#-#-#-#-#-#-#_#-#')
        print(bgp_nei1.group())
        # print(inherits1.group())
        print(bgp_ver1.group())
        print(bgp_state1.group())

    return


def check_ocean_dmvpn(net_connect):
    output = net_connect.send_command('sh dmvpn')

    # filtering only Ocean results
    dmvpn = re.finditer(r'(.*\b124.12\b.*)|(.*\b124.13\b.*)', output)
    # printing results
    print('Ent  Peer NBMA Addr Peer Tunnel Add State  UpDn Tm Attrb')
    for match in dmvpn:
        print(match.group())

    return


def check_mac_add_on_port(mac, net_connect):
    # sending command
    output = net_connect.send_command('sh mac add ad ' + mac)
    # capturing vlanID and int || Po
    vlan_id = re.search(r'(\d+)\s+\S+\s+\bDYNAMIC\b\s+(.*)', output).group(1)
    int = re.search(r'(\d+)\s+\S+\s+\bDYNAMIC\b\s+(.*)', output).group(2)

    # print the vlan ID
    print('==> The mac-add provided belongs to vlan', vlan_id)

    if 'Po' in int:
        print('=> mac-add behind', int)
        po_search = net_connect.send_command('sh etherch summ ' + '| inc' + " " + int)
        # checking if the int is UP
        int_up = re.findall(r'Te\S+|Gi\S+|eth\S+', po_search)
        for j in int_up:
            if 'P' in j:
                # print('=> Interface is up, checking cdp neighbor')
                tmp = re.search(r'(Gi|Te|eth)\d.\d.\d+', j)
                break

        cdp_nei = net_connect.send_command('sh cdp ne ' + str(tmp.group()) + " " + 'de')
        # getting CDP neighbor
        cdp_nei_name = re.search(r'Dev\S+\s\S+\s(.*)', cdp_nei).group(1)
        cdp_nei_ip = re.search(r'IP\sad\S+\s(.*)', cdp_nei).group(1)
        cdp_nei_platform_all = re.search(r'Pla\S+\s\S+\s(\S+)', cdp_nei).group(1)
        cdp_nei_platform = cdp_nei_platform_all.rstrip(',')
        print('The neighbor name is', cdp_nei_name)
        print('The neighbor ip is', cdp_nei_ip)
        print('The neighbor platform is', cdp_nei_platform)

        if 'WS-C3560CX-8PT-S' in cdp_nei_platform:
            print('=> The neighbor is a TL, no more connection')
        else:
            print('*---*.*---*.*---*.*---*.*---*.*---*.*---*')
            print('=> connecting to', cdp_nei_name, 'using', cdp_nei_ip)
            # connecting to nei
            JC = credentials_reconnect(cdp_nei_ip)
            net_connect = ConnectHandler(**JC)
            net_connect.enable()

            # function
            check_mac_add_on_port(mac, net_connect)

    elif 'Gi' in int:
        print('=> mac-add behind', int)
        cdp_nei = net_connect.send_command('sh cdp ne ' + int + " " + 'de')
        if 'Total cdp entries displayed : 0' in cdp_nei or 'Linux' in cdp_nei:
            print('=> No more neighbors')
        else:
            # getting CDP neighbor
            cdp_nei_name = re.search(r'Dev\S+\s\S+\s(.*)', cdp_nei).group(1)
            cdp_nei_ip = re.search(r'IP\sad\S+\s(.*)', cdp_nei).group(1)
            cdp_nei_platform_all = re.search(r'Pla\S+\s\S+\s(\S+)', cdp_nei).group(1)
            cdp_nei_platform = cdp_nei_platform_all.rstrip(',')
            print('The neighbor name is', cdp_nei_name)
            print('The neighbor ip is', cdp_nei_ip)
            print('The neighbor platform is', cdp_nei_platform)

            if 'WS-C3560CX-8PT-S' in cdp_nei_platform:
                print('=> The neighbor is a TL, no more connection')
            else:
                print('=> connecting to', cdp_nei_name, 'using', cdp_nei_ip)
                # connecting to nei
                JC = credentials_reconnect(cdp_nei_ip)
                net_connect = ConnectHandler(**JC)
                net_connect.enable()

                # function
                check_mac_add_on_port(mac, net_connect)


def wlc_get_time(net_connect):
    version = net_connect.send_command("show version")

    if 'Incorrect usage' in version:
        get_time_aireos = net_connect.send_command("show time")
        time = re.search(r'\bTime\b\S+\s(.*)', get_time_aireos).group(1)
        return time
    elif 'Cisco IOS XE Software' in version:
        get_time_ios = net_connect.send_command("show clock")
        # time_ios = re.search(r'[^*].*', get_time_ios).group(1)
        return get_time_ios


def get_wlan_id_wlc(net_connect):
    # checking WLC hardware
    wlc_type = net_connect.send_command('show version')
    if 'Incorrect usage' in wlc_type:
        print('*---.---*.*---.---*.*---.---*')
        print('==> aireOS WLC <==')
        ssid = input(
            "=> If you know the SSID type it. if not, do you want to check all the SSIDs? , (Y) to ONE (N) to all:")
        if ssid in yes_option:
            wlan_summ = net_connect.send_command('show wlan summary')
            wlan_summ_filter = re.findall(r'\/\s(\S+)', wlan_summ)
            # Removing Profile noise in the list
            j = 'SSID'
            while j in wlan_summ_filter:
                wlan_summ_filter.remove(j)
            for item in wlan_summ_filter:
                print(item)

            # asking for SSID
            new_ssid = input('=> Please copy and paste the SSID you want to check:')
            get_wlan_id = net_connect.send_command("show wlan summary")
            wlan_id = re.search(rf"(\d+).*\b{new_ssid}\b", get_wlan_id)
            if not wlan_id:
                print("=> SSID provided wasn't found in WLC, exiting")
                exit(0)
            else:
                return wlan_id.group(1)

        elif ssid != 'n':
            get_wlan_id_ = net_connect.send_command("show wlan summary")
            wlan_id = re.search(rf"(\d+).*\b{ssid}\b", get_wlan_id_)
            if not wlan_id:
                print("=> SSID provided wasn't found in WLC, exiting")
                exit(0)
            else:
                return wlan_id.group(1)

        elif ssid in no_option:
            print("==> Exiting <==")

    else:
        # getting the SSID
        print('*---.---*.*---.---*.*---.---*')
        print('==> IOS WLC <==')
        ssid = input(
            "=> If you know the SSID type it. if not, do you want to check all the SSIDs? , (Y) to ONE (N) to all:")
        if ssid in yes_option:
            wlan_summ = net_connect.send_command('sh wlan summa')
            wlan_summ_filter = re.findall(r'\d+\s+\S+\s+(\S+)', wlan_summ)
            print('==> Printing all the SSIDs')
            # Removing Profile noise in the list
            j = 'Profile'
            while j in wlan_summ_filter:
                wlan_summ_filter.remove(j)
            for item in wlan_summ_filter:
                print(item)

            # asking for SSID
            new_ssid = input('=> Please copy and paste the SSID you want to check:')
            get_wlan_id = net_connect.send_command("sh wlan summa")
            wlan_id = re.search(rf"(\d+).*\b{new_ssid}\b", get_wlan_id)
            if not wlan_id:
                print("=> SSID provided wasn't found in WLC, exiting")
                exit(0)
            else:
                return wlan_id.group(1)

        elif ssid != 'n':
            get_wlan_id_ = net_connect.send_command("sh wlan summa")
            wlan_id = re.search(rf"(\d+).*\b{ssid}\b", get_wlan_id_)
            if not wlan_id:
                print("=> SSID provided wasn't found in WLC, exiting")
                exit(0)
            else:
                return wlan_id.group(1)

        elif ssid in no_option:
            print("==> Exiting <==")


def get_total_clients_ssid_aireos(wlan_id, net_connect):
    print('*---.---*.*---.---*.*---.---*')
    # Getting wlan info
    ssid_id = net_connect.send_command("show wlan " + wlan_id, read_timeout=603)
    # profile_name
    print('=> The SSID name is:=>', re.search(r"Nam\w+\s.SSI\S+\s(.*)", ssid_id).group(1))
    # Status
    status = re.search(r'Name\s.SS\S+.*[\r\n]+\S+\s(.*)', ssid_id).group(1)
    print('=> The SSID status is:=>', status)
    # MAC Filtering
    print('=> The MAC Filtering status:=>',
          re.search(r'Name\s.SS\S+.*[\r\n]+([^\r\n]+)[\r\n]+\S+\s\S+\s(.*)', ssid_id).group(2))
    # MAC Random Filtering
    print('=> The MAC Random Filtering status:=>', re.search(r'Ran\w+\sMA\S+\s\S+\s(.*)', ssid_id).group(1))
    # Broadcast Status
    print('=> The Broadcast status is:=>', re.search(r'Broad\w+\sSS\S+\s(.*)', ssid_id).group(1))
    # Total Clients on SSID
    print('=> Total Number of Active Clients:=>', re.search(r'Numb\w+\s+\S+\sAc\S+\sCl\S+\s(.*)', ssid_id).group(1))
    # Total Random Clients on SSID
    print('=> Total Number of Random Active Clients:=>',
          re.search(r'Numb\w+\s+\S+\sAc\S+\sRa\S+\S+\s\S+\s(.*)', ssid_id).group(1))
    # SSID security
    print('=> SSID Authentication is:=>', re.search(r'802.11\sA\S+\s(.*)', ssid_id).group(1))

    ssid_clients = input(
        "==> do you want to see all the clients connected to this SSID?, (Y) to ONE (N) to all:").lower()
    if ssid_clients in yes_option:
        # Getting all the clients
        ssid_clients = net_connect.send_command('show client wlan ' + wlan_id, read_timeout=603)
        print('=> The local time on WLC is ==>', wlc_get_time(net_connect))
        print('*---.---*.*---.---*.*---.---*')
        print(ssid_clients)

    elif ssid_clients in no_option:
        print('=> Exiting')


def get_total_clients_ssid_9800(wlan_id, net_connect):
    print('*---.---*.*---.---*.*---.---*')
    # Getting wlan info
    ssid_id = net_connect.send_command("sh wlan id " + wlan_id)
    # profile_name
    print('=> The SSID name is:=>', re.search(r"ile\sName\s+.\s(.*)", ssid_id).group(1))
    # Status
    status = re.search(r'Net\w+\sNa\S+\s\S+\s+.\s.*[\r\n]+\S+\s+.\s(.*)', ssid_id).group(1)
    print('=> The SSID status is:=>', status)
    # Broadcast Status
    print('=> The Broadcast status is:=>', re.search(r'Broad\w+\sSS\w+\s+.\s(.*)', ssid_id).group(1))
    # Total Clients on SSID
    print('=> Total Clients on SSID:=>', re.search(r'Number\s\S+\s\S+\s\S+\s+.\s(.*)', ssid_id).group(1))
    # SSID security
    print('=> SSID Authentication is:=>', re.search(r'802.11\sAu\S+\s+.\s(.*)', ssid_id).group(1))

    ssid_clients = input(
        "==> do you want to see all the clients connected to this SSID?, (Y) to ONE (N) to all:").lower()
    if ssid_clients in yes_option:
        # Getting all the clients
        ssid_clients = net_connect.send_command('sho wireless client summary | inc ' + '_' + wlan_id + '_',
                                                read_timeout=603)
        # filtering by mac-add,AP and ID
        print('=> The local time on WLC is ==>', wlc_get_time(net_connect))
        print('*---.---*.*---.---*.*---.---*')
        print(
            'MAC Address    AP Name                                        Type ID   State             Protocol Method     Role')
        print(ssid_clients)

    elif ssid_clients in no_option:
        print('=> Exiting')


def check_wlc_five_two_ghz_ap_status(ap, net_connect):
    # checking is AP is join WLC.
    two_ghz = net_connect.send_command("show ap config 802.11-abgn " + ap, read_timeout=803)
    if 'invalid' in two_ghz:
        print('=> The AP ' + ap + " " + "isn't joined WLC")
        exit(0)
    else:
        print('===> AP is joined WLC, continuing with the process')
        pass

    # getting 5GHZ info
    five_ghz = net_connect.send_command("show ap config 802.11a " + ap, read_timeout=803)
    if 'invalid' in five_ghz:
        print('=> There is no info for 5Ghz radio')
    else:
        pass
    # filtering 5ghz
    five_ghz_tx_level = re.search(r'Curr\w+\sTx\s\w+\s\S+\s\S+\s(.*)', five_ghz).group(1)
    five_ghz_current_channel = re.search(r'Curre\w+\sChan\w+\s\S+\s(.*)', five_ghz).group(1)
    five_ghz_power_config = re.search(r'Tx\sPo\w+\sConf\w+\s+\S+\s(.*)', five_ghz).group(1)
    # filtering 2Ghz
    two_ghz_ghz_tx_level = re.search(r'Curr\w+\sTx\s\w+\s\S+\s\S+\s(.*)', two_ghz).group(1)
    two_ghz_ghz_current_channel = re.search(r'Curre\w+\sChan\w+\s\S+\s(.*)', two_ghz).group(1)
    two_ghz_ghz_power_config = re.search(r'Tx\sPo\w+\sConf\w+\s+\S+\s(.*)', two_ghz).group(1)

    # printing the 2.5 and 5Ghz
    print('==> Getting 2.4Ghz info for ' + ap)
    print('=> AP TX_Level config is:', two_ghz_ghz_power_config)
    print('=> AP TX_Level is:', two_ghz_ghz_tx_level)
    print('=> AP Current Channel is:', two_ghz_ghz_current_channel)
    print('*---*.*---*.*---*.*---*.*---*.*---*.')
    print('==> Getting 5Ghz info for ' + ap)
    print('=> AP TX_Level config is:', five_ghz_power_config)
    print('=> AP TX_Level is:', five_ghz_tx_level)
    print('=> AP Current Channel is:', five_ghz_current_channel)
    print('*---*.*---*.*---*.*---*.*---*.*---*.')


def set_wlc_qos(wlan_id, net_connect):
    print(
        '==>*** Changing WLAN parameters while it is enabled will cause the WLAN to be momentarily disabled and radio reset thus may result in loss of connection ***<==\n')

    # asking for rate vlue and validating input of 3,4 integers or 0
    try:
        rate = int(input(
            "Please enter the QoS value(kbps), it will need to be 1, 3 or 4 Integers values starting from 0 to 6, 0 disable QoS: "))
    except ValueError:
        print("Sorry, I didn't understand that.")
        exit(0)

    if re.match('^[0-6]{1}[0-9]{3}$|^[0-6]{1}[0-9]{2}$|^0{1}', str(rate)):
        pass
        # print("Entry is valid")

    else:
        print("Sorry, your entry is not correct, we need 4 integers from 0 to 6.")
        exit(0)

    # Getting WLAN info
    wlan_info = net_connect.send_command("show wlan " + wlan_id)
    wlan_name = re.search(r'Profi\w+\s\S+\s(.*)', wlan_info).group(1)

    # commands for QoS
    # Average Data Rate
    override_rate_limit_down_cmd = 'config wlan override-rate-limit ' + wlan_id + ' average-data-rate per-client downstream ' + str(
        rate)
    override_rate_limit_up_cmd = 'config wlan override-rate-limit ' + wlan_id + ' average-data-rate per-client upstream ' + str(
        rate)
    # Average Real-Time Rate
    average_realtime_rate_down_cmd = 'config wlan override-rate-limit ' + wlan_id + ' average-realtime-rate per-client downstream ' + str(
        rate)
    average_realtime_rate_up_cmd = 'config wlan override-rate-limit ' + wlan_id + ' average-realtime-rate per-client upstream ' + str(
        rate)
    # Burst Data Rate
    burst_data_rate_down_cmd = 'config wlan override-rate-limit ' + wlan_id + ' burst-data-rate per-client downstream ' + str(
        rate)
    burst_data_rate_up_cmd = 'config wlan override-rate-limit ' + wlan_id + ' burst-data-rate per-client upstream ' + str(
        rate)
    # Burst Real-Time Rate
    burst_realtime_rate_down_cmd = 'config wlan override-rate-limit ' + wlan_id + ' burst-realtime-rate per-client downstream ' + str(
        rate)
    burst_realtime_rate_up_cmd = 'config wlan override-rate-limit ' + wlan_id + ' burst-realtime-rate per-client upstream ' + str(
        rate)
    # commands end #

    # disabling the Wlan
    disable_cmd = 'config wlan disable ' + wlan_id
    disable_cmd_send = net_connect.send_command(disable_cmd)
    # if disable_cmd_send:
    #    pass
    #    # print(' -> WLAN ' + wlan_name + ' has been disabled')
    # else:
    #    print(' -> Something went wrong disabling the WLAN ' + wlan_name)

    # Sending commands
    # sending override_rate_limit
    override_rate_limit_down_cmd_send = net_connect.send_command(override_rate_limit_down_cmd)
    # if not override_rate_limit_down_cmd_send:
    #    print('=> Something went wrong with override_rate_limit_down')
    override_rate_limit_up_cmd_send = net_connect.send_command(override_rate_limit_up_cmd)
    # if not override_rate_limit_up_cmd_send:
    #    print('=> Something went wrong with override_rate_limit_up')

    # Sending average_realtime_rate
    average_realtime_rate_down_cmd_send = net_connect.send_command(average_realtime_rate_down_cmd)
    # if not average_realtime_rate_down_cmd_send:
    #    print('=> Something went wrong with average_realtime_rate_down')
    average_realtime_rate_up_cmd_send = net_connect.send_command(average_realtime_rate_up_cmd)
    # if not average_realtime_rate_up_cmd_send:
    #    print('=> Something went wrong with average_realtime_rate_up')

    # Sending burst_data_rate
    burst_data_rate_down_cmd_send = net_connect.send_command(burst_data_rate_down_cmd)
    # if not burst_data_rate_down_cmd_send:
    #    print('=> Something went wrong with burst_data_rate_down')
    burst_data_rate_up_cmd_send = net_connect.send_command(burst_data_rate_up_cmd)
    # if not burst_data_rate_up_cmd_send:
    #    print('=> Something went wrong with burst_data_rate_up')

    # Sending burst_realtime_rate
    burst_realtime_rate_down_cmd_send = net_connect.send_command(burst_realtime_rate_down_cmd)
    # if not burst_realtime_rate_down_cmd_send:
    #    print('=> Something went wrong with burst_realtime_rate_down')
    burst_realtime_rate_up_cmd_send = net_connect.send_command(burst_realtime_rate_up_cmd)
    # if not burst_realtime_rate_up_cmd_send:
    #    print('=> Something went wrong with burst_realtime_rate_up')

    # Enabling WLAN
    enable_cmd = 'config wlan enable ' + wlan_id
    enable_cmd_send = net_connect.send_command(enable_cmd)
    if enable_cmd_send:
        print('==> QoS parameters has been executed <==')
        # print(' -> WLAN ' + wlan_name + ' has been enabled')
    else:
        print(' -> Something went wrong enabling the WLAN ' + wlan_name + '\n')


def set_wlc_ap_tx_power(ap, net_connect):
    # checking AP 2.4 and 5Ghz radio status
    check_wlc_five_two_ghz_ap_status(ap, net_connect)

    tx_power_change = input(
        "==> after checking the results, do you want to change TX_POWER?, (Y) to ONE (N) to all:").lower()
    if tx_power_change in yes_option:
        tx_power_value = input("What's the TX value? being 1 the highest and 7 the lowest: ").upper()
        # disabling 5Ghz radio before the tx_power change
        five_tx_power_modification_command = 'config 802.11a disable ' + ap
        five_tx_power_modification_command_send = net_connect.send_command(five_tx_power_modification_command,
                                                                           read_timeout=803)
        if 'invalid' in five_tx_power_modification_command_send:
            print('=> Disabled radios command failed')
            exit(0)
        else:
            pass
            # print('=> 5Ghz Radios disabled successfully')

        # disabling 2.5Ghz radio before the tx_power change
        two_tx_power_modification_command = 'config 802.11-abgn disable ' + ap
        two_tx_power_modification_command_send = net_connect.send_command(two_tx_power_modification_command,
                                                                          read_timeout=803)
        if 'invalid' in two_tx_power_modification_command_send:
            print('=> Disabled radios command failed')
            exit(0)
        else:
            pass
            # print('=> 2.5Ghz Radios disabled successfully')

        # enabling manual mode client-serving for 2.4Ghz
        two_tx_manualMode_modification_command = 'config 802.11-abgn role ' + ap + " " + 'manual client-serving'
        two_tx_manualMode_modification_command_send = net_connect.send_command(two_tx_manualMode_modification_command,
                                                                               read_timeout=803)
        if 'invalid' in two_tx_manualMode_modification_command_send:
            print('=> Manual Mode radios command failed')
            exit(0)
        else:
            pass
            # print('=> 2.5Ghz Radios disabled successfully')

        # Changing TX_POWER value 5Ghz
        five_tx_power_value_command = 'config 802.11a txPower ap ' + ap + " " + tx_power_value
        five_tx_power_value_command_send = net_connect.send_command(five_tx_power_value_command, read_timeout=803)
        if 'invalid' in five_tx_power_value_command_send:
            print('=> TX_POWER command failed')
            exit(0)
        else:
            print('=> 5Ghz Radio TX_POWER changed successfully')
        # Changing TX_POWER value 2.4Ghz
        two_tx_power_value_command = 'config 802.11-abgn txPower ap ' + ap + " " + tx_power_value
        two_tx_power_value_command_send = net_connect.send_command(two_tx_power_value_command, read_timeout=803)
        if 'invalid' in two_tx_power_value_command_send:
            print('=> TX_POWER command failed')
            exit(0)
        else:
            print('=> 2.4Ghz Radio TX_POWER changed successfully')

        # enabling radios after the change
        enable_five_tx_power_modification_command = 'config 802.11a enable ' + ap
        enable_five_tx_power_modification_command_send = net_connect.send_command(
            enable_five_tx_power_modification_command, read_timeout=803)
        if 'invalid' in enable_five_tx_power_modification_command_send:
            print('=> Enabling radios command failed')
            exit(0)
        else:
            pass
            # print('=> 5Ghz Radios enabled successfully')
        # enabling 2.5Ghz radio before the tx_power change
        enable_two_tx_power_modification_command = 'config 802.11-abgn enable ' + ap
        enable_two_tx_power_modification_command_send = net_connect.send_command(
            enable_two_tx_power_modification_command, read_timeout=803)
        if 'invalid' in enable_two_tx_power_modification_command_send:
            print('=> Enabling radios command failed')
            exit(0)
        else:
            pass
            # print('=> 2.5Ghz Radios enabled successfully')
        print('*---*.*---*.*---*.*---*.*---*.*---*.')
        print('*---*.*---*.*---*.*---*.*---*.*---*.')

        # showing results after the change
        # checking AP 2.4 and 5Ghz radio status
        print('=> Showing results after the change')
        check_wlc_five_two_ghz_ap_status(ap, net_connect)

        print('=> Exiting <=')
        exit(0)

    elif tx_power_change in no_option:
        print("==> No TX_POWER commands executed <==")
        pass

    ### setting default values to AP TX_POWER ###
    tx_power_default = input("==> Do you want to default the TX_POWER value?, (Y) to ONE (N) to all:").lower()
    if tx_power_default in yes_option:
        # disabling 5Ghz radio before the tx_power change
        five_tx_power_modification_command = 'config 802.11a disable ' + ap
        five_tx_power_modification_command_send = net_connect.send_command(five_tx_power_modification_command,
                                                                           read_timeout=803)
        if 'invalid' in five_tx_power_modification_command_send:
            print('=> Disabled radios command failed')
            exit(0)
        else:
            pass
            # print('=> 5Ghz Radios disabled successfully')
        # disabling 2.4Ghz radio before the tx_power change
        two_tx_power_modification_command = 'config 802.11-abgn disable ' + ap
        two_tx_power_modification_command_send = net_connect.send_command(two_tx_power_modification_command,
                                                                          read_timeout=803)
        if 'invalid' in two_tx_power_modification_command_send:
            print('=> Disabled radios command failed')
            exit(0)
        else:
            pass
            # print('=> 2.5Ghz Radios disabled successfully')

        # enabling manual mode client-serving for 2.4Ghz
        two_tx_manualMode_modification_command = 'config 802.11-abgn role ' + ap + " " + 'manual client-serving'
        two_tx_manualMode_modification_command_send = net_connect.send_command(
            two_tx_manualMode_modification_command, read_timeout=803)
        if 'invalid' in two_tx_manualMode_modification_command_send:
            print('=> Manual Mode radios command failed')
            exit(0)
        else:
            pass
            # print('=> 2.5Ghz Radios disabled successfully')

        # executing default values
        # Changing TX_POWER value 5Ghz
        five_tx_power_value_command = 'config 802.11a txPower ap ' + ap + " " + 'global'
        five_tx_power_value_command_send = net_connect.send_command(five_tx_power_value_command,
                                                                    read_timeout=803)
        if 'invalid' in five_tx_power_value_command_send:
            print('=> TX_POWER default command failed')
            exit(0)
        else:
            print('=> 5Ghz default Radio TX_POWER changed successfully')
        # Changing TX_POWER value 2.4Ghz
        two_tx_power_value_command = 'config 802.11-abgn txPower ap ' + ap + " " + 'global'
        two_tx_power_value_command_send = net_connect.send_command(two_tx_power_value_command, read_timeout=803)
        if 'invalid' in two_tx_power_value_command_send:
            print('=> TX_POWER default command failed')
            exit(0)
        else:
            print('=> 2.4Ghz Default Radio TX_POWER changed successfully')

        # enabling 2.5Ghz radio before the tx_power change
        enable_five_tx_power_modification_command = 'config 802.11a enable ' + ap
        enable_five_tx_power_modification_command_send = net_connect.send_command(
            enable_five_tx_power_modification_command, read_timeout=803)
        if 'invalid' in enable_five_tx_power_modification_command_send:
            print('=> Enabling radios command failed')
            exit(0)
        else:
            pass
            # print('=> 5Ghz Radios enabled successfully')
        # enabling 2.5Ghz radio before the tx_power change
        enable_two_tx_power_modification_command = 'config 802.11-abgn enable ' + ap
        enable_two_tx_power_modification_command_send = net_connect.send_command(
            enable_two_tx_power_modification_command, read_timeout=803)
        if 'invalid' in enable_two_tx_power_modification_command_send:
            print('=> Enabling radios command failed')
            exit(0)
        else:
            pass
            # print('=> 2.5Ghz Radios enabled successfully')

        # enabling auto mode for 2.4Ghz
        two_tx_manualMode_modification_command = 'config 802.11-abgn role ' + ap + " " + 'auto'
        two_tx_manualMode_modification_command_send = net_connect.send_command(
            two_tx_manualMode_modification_command, read_timeout=803)
        if 'invalid' in two_tx_manualMode_modification_command_send:
            print('=> Manual Mode radios command failed')
            exit(0)
        elif 'Channel and/or radio role selection may be sub-optimal' in two_tx_manualMode_modification_command_send:
            print('Auto mode command was selected')
        print('*---*.*---*.*---*.*---*.*---*.*---*.')
        print('*---*.*---*.*---*.*---*.*---*.*---*.')

        # showing results after the change
        # checking AP 2.4 and 5Ghz radio status
        print('=> Showing results after the change')
        check_wlc_five_two_ghz_ap_status(ap, net_connect)

    elif tx_power_default in no_option:
        print("==> No Default TX_POWER applied <==")
        exit(0)


def wlc_client_count_by_ap_9800(client_count, net_connect):
    ap_summ = net_connect.send_command("sh ap summary sort descending client-count", read_timeout=603)
    ap_total_summ = net_connect.send_command("sh ap summary | inc Number", read_timeout=603)
    ap_filter = re.findall(r'(\w+\S+)\s+\S+\s+(\d+)\s+', ap_summ)
    ap_total = re.search(r'Num\w+\s\S+\s\S+\s(\w+)', ap_total_summ)
    print('=> The local time on WLC is ==>', wlc_get_time(net_connect))
    print('=> Total Numbers of APs joined WLC =', ap_total.group(1))
    for item in ap_filter:
        if int(item[1]) >= client_count:
            print(item)


def wlc_client_count_by_ap(client_count, net_connect):
    ap_summ = net_connect.send_command("show ap summ", read_timeout=603)
    ap_total = re.search(r'Num\w+\s\S+\s\S+\s(\w+)', ap_summ)
    # filtering results
    ap_filter = re.findall(r'(\S+)\s+\d+\s+\S+\s+\S+\s+\w+\s\S+\s+\w+\s+\d+\S+\s+(\d+)', ap_summ)
    # print('==> Date =', get_time_date()[0], '=> Time =', get_time_date()[1])
    print('=> The local time on WLC is ==>', wlc_get_time(net_connect))
    print('=> Total Numbers of APs joined WLC =', ap_total.group(1))
    ### result = [tup for tup in ap_filter if int(tup[1]) > 21]
    for item in ap_filter:
        if int(item[1]) >= client_count:
            print(item)
        # elif
        #   print('=> There were no APs with more than ' + str(client_count) + " " + 'clients joined WLC')


def wlc_utils_ap(ap_name, net_connect):
    version = net_connect.send_command("show version")

    if 'Incorrect usage' in version:
        print('==> Using aireOS WLC <==')
        # validation if you enter a mac-add
        if re.match(r'([0-9A-Fa-f]{2}[:-]){5}([0-9A-Fa-f]{2})', ap_name):
            if_mac_add = net_connect.send_command('show ap config general ' + ap_name, read_timeout=801)
            ap_name = re.search(r'Cisco\sAP\sNa\S+\s(.*)', if_mac_add).group(1)
            print('=> Because you provided a mac-add, this is the current AP name:', ap_name, '\n')

        # checking AP neighbor
        cdp = net_connect.send_command('show ap cdp neighbors ap-name ' + ap_name)
        inv = net_connect.send_command('show ap inventory ' + ap_name)
        if 'Invalid AP name specified' in cdp:
            print("==> AP isn't joined WLC, exiting")
            exit(0)
        else:
            print('=> Local WLC time is:', wlc_get_time(net_connect))
            print('====> Showing CDP neighbor details <====')
            print('=> The AP name is:', ap_name)
            print('=> The neighbor name is:',
                  re.search(r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\s+(\S+)\s+(.*)', cdp).group(1))
            print('=> The Neighbor Interface is:',
                  re.search(r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\s+(\S+)\s+(.*)', cdp).group(2))
            print('=> The neighbor IP is:', re.search(r'IP\sadd\S+\s(.*)', cdp).group(1))
            print('====> Inventory Info <====')
            print('=> The AP description is:', re.search(r'DE\S+\s(.*)', inv).group(1))
            inv_model = re.search(r'PID\S+\s(\S+)', inv).group(1).rstrip('.')
            print('=> The AP model is:', inv_model)
            print('=> The AP SN is:', re.search(r'SN.\s(.*)', inv).group(1))
            print('***-***.***-***.***-***.***-***.***-***')

            ap_options = input(
                '==> Please select one of the following options:\n => Select 1 to Reboot AP\n => Select 2 to Disable AP\n => Select 3 to Enable AP\n => Select 4 AP LED Flash \n => Select 5 to Rename AP \n => Select 6 to Change TX_POWER \n => Select 7 to Check if the AP is joined WLC \n:').lower()

            if '1' in ap_options:
                print('=> Rebooting AP =>', ap_name)
                reset_command = 'config ap reset ' + ap_name
                reset = net_connect.send_command(reset_command,
                                                 expect_string=r'Would you like to reset ' + ap_name + " " + '?')
                try:
                    reset += net_connect.send_command('y', expect_string=r'>')

                except:
                    raise

                # checking if the AP went down
                checking_ap = net_connect.send_command('show ap cdp neighbors ap-name ' + ap_name)
                if 'Invalid AP name specified' in checking_ap:
                    print('=> AP', ap_name, 'was rebooted')
                else:
                    print('=> AP still UP')
                exit(0)

            elif '2' in ap_options:
                print('=> Disabling AP =>', ap_name)
                disable_command = 'config ap disable ' + ap_name
                disable = net_connect.send_command(disable_command)
                # checking if the AP went down
                ap_status_general = net_connect.send_command('show ap config general ' + ap_name)
                ap_status = re.search(r'Adminis\S+\sSta\w+\s\S+\s(.*)', ap_status_general).group(1)
                if 'ADMIN_DISABLED' in ap_status:
                    print('=> AP has been disabled')
                else:
                    print('=> AP still enable')
                exit(0)

            elif '3' in ap_options:
                print('=> Enabling AP =>', ap_name)
                enable_command = 'config ap enable ' + ap_name
                enable = net_connect.send_command(enable_command)
                # checking if the AP is enabled
                ap_status_general = net_connect.send_command('show ap config general ' + ap_name)
                ap_status = re.search(r'Adminis\S+\sSta\w+\s\S+\s(.*)', ap_status_general).group(1)
                if 'ADMIN_ENABLED' in ap_status:
                    print('=> AP has been enabled')
                else:
                    print('=> AP still disabled')
                exit(0)

            elif '4' in ap_options:
                print('=> Enabling AP Blinking LED =>', ap_name)
                blink_led_input = int(input('=> Please provide the time(in seconds) to flash the LED on the AP: '))
                blink_led = 'config ap led-state flash ' + str(blink_led_input) + " " + ap_name
                blink_send = net_connect.send_command(blink_led)
                print('=> LED Blink command has been sent to ' + ap_name + " " + 'for ' + str(
                    blink_led_input) + " " + 'seconds')

            elif '5' in ap_options:
                print('=> Renaming =>', ap_name)
                new_ap_name = (input('=> Please provide the new AP name: '))
                rename_ap = 'config ap name ' + new_ap_name + " " + ap_name
                rename_ap_sent = net_connect.send_command(rename_ap)
                if not 'invalid' in rename_ap_sent:
                    print('=> AP has been renamed')
                else:
                    print('=> AP rename fail')

            elif '6' in ap_options:
                print('=> Changing TX_LEVEL', ap_name)
                # calling function
                set_wlc_ap_tx_power(ap_name, net_connect)

            elif '7' in ap_name:
                print('=> Checking if the AP is joined WLC', ap_name)
                # calling function
                get_ios_wlc_ap(ap_name, net_connect)

            else:
                print('=> Wrong option selected')
                exit(0)

    elif 'Cisco IOS XE Software' in version:
        print('==> Using IOS WLC (9800) <==')
        # checking AP neighbor
        cdp = net_connect.send_command('sh ap name ' + ap_name + " " + 'cdp neighbors')
        inv = net_connect.send_command('sh ap name ' + ap_name + " " + 'inven')
        if not cdp:
            print("==> AP isn't joined WLC, exiting")
            exit(0)
        else:
            print('=> Local WLC time is:', wlc_get_time(net_connect))
            print('====> Showing CDP neighbor details <====')
            print('=> The AP name is:', ap_name)
            print('=> The neighbor name is:',
                  re.search(r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\s+(\S+)\s+(\S+)\s+(.*)', cdp).group(1))
            print('=> The Neighbor Interface is:',
                  re.search(r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\s+(\S+)\s+(\S+)\s+(\S+)', cdp).group(3))
            print('=> The neighbor IP is:',
                  re.search(r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\s+(\S+)\s+(\S+)\s+(\S+)', cdp).group(2))
            print('====> Inventory Info <====')
            print('=> The AP description is:', re.search(r'DE\S+\s(.*)', inv).group(1))
            inv_model = re.search(r'PID\S+\s(\S+)', inv).group(1).rstrip(',')
            print('=> The AP model is:', inv_model)
            print('=> The AP SN is:', re.search(r'SN.\s(.*)', inv).group(1))
            print('***-***.***-***.***-***.***-***.***-***')

            ap_options = input(
                '==> Please select one of the following options:\n => Select 1 to Reboot AP\n => Select 2 to Disable AP\n => Select 3 to Enable AP\n => Select 4 AP LED Flash \n => Select 5 to Rename AP \n:').lower()

            if '1' in ap_options:
                print('=> Rebooting AP =>', ap_name)
                reset_command = 'ap name ' + ap_name + " " + 'reset'
                reset = net_connect.send_command(reset_command)

                # checking if the AP went down
                checking_ap = net_connect.send_command('sh ap name ' + ap_name + " " + 'cdp neighbors')
                if not checking_ap:
                    print('=> AP', ap_name, 'was rebooted')
                else:
                    print('=> AP still UP')
                exit(0)

            elif '2' in ap_options:
                print('=> Disabling AP =>', ap_name)
                disable_command = 'ap name ' + ap_name + " " + 'shutdown'
                disable = net_connect.send_command(disable_command)
                # checking if the AP went down
                ap_status_general = net_connect.send_command('show ap name ' + ap_name + " " + 'config general')
                ap_status = re.search(r'Admi\S+\sSt\S+\s+.\s(.*)', ap_status_general).group(1)
                if 'Disabled' in ap_status:
                    print('=> AP has been disabled')
                else:
                    print('=> AP still enable')
                exit(0)

            elif '3' in ap_options:
                print('=> Enabling AP =>', ap_name)
                enable_command = 'ap name ' + ap_name + " " + 'no shutdown'
                enable = net_connect.send_command(enable_command)
                # checking if the AP is enabled
                ap_status_general = net_connect.send_command('show ap name ' + ap_name + " " + 'config general')
                ap_status = re.search(r'Admi\S+\sSt\S+\s+.\s(.*)', ap_status_general).group(1)
                if 'Enabled' in ap_status:
                    print('=> AP has been enabled')
                else:
                    print('=> AP still disabled')
                exit(0)

            elif '4' in ap_options:
                print('=> Enabling AP Blinking LED =>', ap_name)
                blink_led_input = int(input('=> Please provide the time(in seconds) to flash the LED on the AP: '))
                blink_led = 'ap name ' + ap_name + " " + 'led flash start duration' + str(blink_led_input)
                blink_send = net_connect.send_command(blink_led)
                print('=> LED Blink command has been sent to ' + ap_name + " " + 'for ' + str(
                    blink_led_input) + " " + 'seconds')

            elif '5' in ap_options:
                print('=> Renaming =>', ap_name)
                new_ap_name = (input('=> Please provide the new AP name: '))
                rename_ap = 'ap name ' + ap_name + " " + 'name ' + new_ap_name
                rename_ap_sent = net_connect.send_command(rename_ap)
                if not 'invalid' in rename_ap_sent:
                    print('=> AP has been renamed')
                else:
                    print('=> AP rename fail')

            else:
                print('=> Wrong option selected')
                exit(0)


def getting_model(net_connect):
    dev_model = net_connect.send_command('show version | inc Model', read_timeout=603)
    # dev_model = cli('show version | inc Model')
    mod_detail = re.search(r'Mo\S+\s[N|n]um\S+\s+.\s(.*)', dev_model).group(1)
    # print('=> Model = ' + mod_detail)
    return mod_detail


def connect_wlc(isIP):
    if isIP:
        print('*---*-*---*-*---*-*---*-*---*')
        IP = isIP
        print('==> Using CCL credentials')
        # USERNAME = input("What's the username: ")
        # PASS = input("What's the password: ")
        # PASS = getpass.getpass()
        # Connection to WLC
        JC = {
            'device_type': 'cisco_wlc_ssh',
            'ip': IP,
            'username': 'ccl',
            'password': 'N@v!gaT!nG~',
            "timeout": 603,
        }
        return JC

    else:
        pass

    IP = input("Give me the device IP: ")
    USERNAME = input("What's the username: ")
    # PASS = input("What's the password: ")
    PASS = getpass.getpass()

    # Connection to WLC

    JC = {
        'device_type': 'cisco_wlc_ssh',
        'ip': IP,
        'username': USERNAME,
        'password': PASS,
    }

    return JC


def set_wlc_policy_tag_9800(ap_name, net_connect):
    output = net_connect.send_command("sh ap name " + ap_name + " " + 'config general')
    tag_policy = re.search(r'Policy\sT\w+\sNa\w+\s+.\s(.*)', output).group(1)
    mac = re.search(r'MAC\sAdd\w+\s+.\s(\d\S+|\w\S+)', output).group(1)

    print('mac-add=>', mac)
    # Changing TAG-Policy
    config_commands = ['config t', 'ap' + " " + mac, 'policy-tag XIC']
    output = net_connect.send_config_set(config_commands)
    if 'Associating policy-tag will cause associated AP to reconnect' in output:
        print('Command executed successfully')


def get_wlc_mimosa_check(mac, net_connect):
    output = net_connect.send_command("show ap config general" + " " + mac)
    if 'Cisco AP name is invalid' in output:
        print("=> AP with mac-add = " + mac + " " + "isn't joined WLC")
        print('*---*-*---*-*---*-*---*')

    else:
        ap_name = re.search(r'AP\sNa\S+\s(.*)', output)
        ipv4_bulk = re.search(r'IP\sAdd\w+\sConfiguration.*[\r\n]+([^\r\n]+)', output).group(1)
        ipv4 = re.search(r'IP\sAddre\w+\S+\s(.*)', ipv4_bulk)
        group_name = re.search(r'Group\sNam\S+\s(.*)', output)
        return ap_name.group(1), ipv4.group(1), group_name.group(1)


# function to get specific details from Wireless Client
def wlc_aireos_client(mac, net_connect):
    print('==> Date =', get_time_date()[0], '=> Time =', get_time_date()[1])
    output = net_connect.send_command("show  client detail" + " " + mac)
    x = ' '
    print('*---*-*---*-*---*-*---*-*---*')
    print('=> Client mac-add =', mac)
    print('Client Username =>', x * 22, re.search(r'Clie\S+\sUsern\S+\s\S+\s(.*)', output).group(1))
    print('Client Webauth Username =>', x * 14, re.search(r'Clie\S+\sWe\S+\s\S+\s\S+\s(.*)', output).group(1))
    print('Hostname =>', x * 29, re.search(r'Host\w+.\s\S+\s(.*)', output).group(1))
    print('Device Type =>', x * 26, re.search(r'Dev\S+\s\S+.\S+\s(.*)', output).group(1))
    print('Connected to =>', x * 25, re.search(r'AP\sNa\S+\s(\S+)', output).group(1))
    print('Client State =>', x * 25, re.search(r'Cli\w+\sSt\S+\s(.*)', output).group(1))
    print('Wireless LAN Id =>', x * 22, re.search(r'Wire\w+\sL\w+\sI\S+\s(.*)', output).group(1))
    print('Wireless LAN Network Name (SSID) =>', x * 5,
          re.search(r'Wire\w+\sL\w+\sNe\S+\s\w+\s\S+\s(.*)', output).group(1))
    print('Wireless LAN Profile Name =>', x * 12, re.search(r'Wire\w+\sL\w+\sPr\w+\s\S+\s(.*)', output).group(1))
    print('Connected For =>', x * 24, re.search(r'Conn\w+\sF\S+\s\S+\s(.*)', output).group(1))
    print('BSSID =>', x * 32, re.search(r'BSS\S+\s(\S+)', output).group(1))
    print('Channel =>', x * 30, re.search(r'Chan\S+\s(\S+)', output).group(1))
    print('IP Address =>', x * 27, re.search(r'IP\sAdd\S+\s(.*)', output).group(1))
    print('Gateway Address =>', x * 22, re.search(r'Gate\S+\s\S+\s(.*)', output).group(1))
    print('Netmask =>', x * 30, re.search(r'Netm\S+\s(.*)', output).group(1))
    print('Association Id =>', x * 23, re.search(r'Asso\S+\s\S+\s(.*)', output).group(1))
    print('Authentication Algorithm =>', x * 13, re.search(r'Authenticatio\S+\s\S+\s(.*)', output).group(1))
    print('Reason Code =>', x * 26, re.search(r'Reas\w+\s\S+\s(.*)', output).group(1))
    print('==> QoS <==')
    print('QoS Level =>', x * 28, re.search(r'QoS\s\S+\s(.*)', output).group(1))
    print('Avg data Rate =>', x * 24, re.search(r'Avg\sda\S+\s\S+\s(.*)', output).group(1))
    print('Burst data Rate =>', x * 22, re.search(r'Burst\sda\S+\s\S+\s(.*)', output).group(1))
    print('Avg Real time data Rate =>', x * 14, re.search(r'Avg\sRe\S+\s\S+\s\S+\s\S+\s(.*)', output).group(1))
    print('Burst Real Time data Rate =>', x * 12, re.search(r'Burst\sda\S+\s\S+\s(.*)', output).group(1))
    print('Avg Uplink data Rate =>', x * 17, re.search(r'Avg\sUp\w+\sda\w+\s\S+\s(.*)', output).group(1))
    print('Burst Uplink data Rate =>', x * 15, re.search(r'Bur\w+\sUp\w+\sda\w+\s\S+\s(.*)', output).group(1))
    print('Avg Uplink Real time data Rate =>', x * 7,
          re.search(r'Avg\sUp\w+\sRe\w+\st\w+\sda\w+\sRa\S+\s(.*)', output).group(1))
    print('Burst Uplink Real Time data Rate =>', x * 5,
          re.search(r'Bur\w+\sUp\w+\sRe\w+\s\w+\s\w+\s\S+\s(.*)', output).group(1))
    print('Supported Rates =>', x * 22, re.search(r'Supp\w+\sRa\S+\s(.*)', output).group(1))
    print('Mobility State =>', x * 23, re.search(r'Mobi\w+\sSta\S+\s(.*)', output).group(1))
    print('Mobility Move Count =>', x * 18, re.search(r'Mobil\w+\sMo\w+\sC\S+\s(.*)', output).group(1))
    print('Audit Session ID =>', x * 21, re.search(r'Audit\sS\w+\s\S+\s(.*)', output).group(1))
    print('Interface =>', x * 28, re.search(r'Interface\S+\s(.*)', output).group(1))
    print('VLAN =>', x * 33, re.search(r'Interface.*[\r\n]+(\w+\S+\s(.*))', output).group(2))
    print('Access VLAN =>', x * 26, re.search(r'Acce\w+\sVLA\S+\s(.*)', output).group(1))
    print('Local Bridging VLAN =>', x * 18, re.search(r'Local\sBr\w+\s\S+\s(.*)', output).group(1))
    print('==> Client Statistics <==')
    print('Number of Bytes Received =>', x * 13,
          re.search(r'Client Statistics.[\r\n]+(\s+\w+\s\S+\s\S+\s\S+\s(.*))', output).group(2))
    print('Number of Bytes Sent =>', x * 17,
          re.search(r'Number of Bytes Received.*[\r\n]+(\s+\S+\s\S+\s\S+\s\S+\s(.*))', output).group(2))
    print('Total Number of Bytes Sent =>', x * 11, re.search(r'Total\s\S+\s\S+\s\S+\sSe\S+\s(.*)', output).group(1))
    print('Total Number of Bytes Recv =>', x * 11, re.search(r'Total\s\S+\s\S+\s\S+\sRe\S+\s(.*)', output).group(1))
    print('Number of Bytes Sent (last 90s) =>', x * 6,
          re.search(r'Total Number of Bytes Recv.*[\r\n]+(\s+\S+\s\S+\s\S+\s\S+\s\S+\s\S+\s(.*))', output).group(2))
    print('Number of Bytes Recv (last 90s) =>', x * 6,
          re.search(r'Total Number of Bytes Recv.*[\r\n]+([^\r\n]+)(\s+\S+\s\S+\s\S+\s\S+\s\S+\s\S+\s(.*))',
                    output).group(3))
    print('Number of Packets Received =>', x * 11, re.search(r'Number\s\S+\sPa\S+\sRe\S+\s(.*)', output).group(1))
    print('Number of Packets Sent =>', x * 15, re.search(r'Number\s\S+\sPa\S+\sSen\S+\s(.*)', output).group(1))
    print('Number of Data Retries =>', x * 15, re.search(r'Data\sRe\S+\s(.*)', output).group(1))
    print('Radio Signal Strength Indicator =>', x * 6, re.search(r'Radi\S+\s\S+\s\S+\s\S+\s(.*)', output).group(1))
    print('Signal to Noise Ratio =>', x * 16, re.search(r'Sig\w+\s\S+\sNo\S+\s\S+\s(.*)', output).group(1))
    print(re.search(
        r'Clie\w+\sCapa\S+[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)',
        output).group())


# function to get all the Client details
def wlc_aireos_client_details(mac, net_connect):
    print('==> Date =', get_time_date()[0], '=> Time =', get_time_date()[1])
    output = net_connect.send_command("show  client detail" + " " + mac)
    print(output)


def wlc_clients_associated_ap_details(ap_name, net_connect):
    output = net_connect.send_command("show ap config general" + " " + ap_name, read_timeout=703)
    five_radio = net_connect.send_command("show ap config 802.11a " + ap_name, read_timeout=703)
    two_radio = net_connect.send_command("show ap config 802.11-abgn " + ap_name, read_timeout=703)

    print('*---*-*---*-*---*-*---*-*---*')
    ap_name = re.search(r'AP\sNa\w+\S+\s(.*)', output)
    ap_mac = re.search(r'MAC\sAd\w+\S+\s([0-9a-f].*)', output)
    ipv4_mode = re.search(r'Addr\w+\sCo\S+\s(.*)', output)
    ipv4_bulk = re.search(r'IP\sAdd\w+\sConfiguration.*[\r\n]+([^\r\n]+)', output).group(1)
    ipv4 = re.search(r'IP\sAddre\w+\S+\s(.*)', ipv4_bulk)
    netmask = re.search(r'NetMas\S+\s(.*)', output)
    gw = re.search(r'Gatew\S+\s\w+\s\S+\s(.*)', output)
    capwap_path_mtu = re.search(r'CAPW\S+\s\S+\s\S+\s(.*)', output)
    location = re.search(r'Locati\S+\s(.*)', output)
    group_name = re.search(r'Group\sNam\S+\s(.*)', output)
    admin_state = re.search(r'Adminis\S+\s\S+\s\S+\s(.*)', output)
    operation_mode = re.search(r'Opera\w+\s\S+\s\S+\s(.*)', output)
    ap_mode = re.search(r'AP\sMode\s\S+\s(.*)', output)
    soft_ver = re.search(r'S\/W\s+\S+\s\S+\s(.*)', output)
    ap_model = re.search(r'AP\sModel\S+\s(.*)', output)
    ap_image = re.search(r'AP\sIma\S+\s(.*)', output)
    sn = re.search(r'Seri\S+\s\S+\s(.*)', output)
    ap_uptime = re.search(r'AP\sUp\s\S+\s(.*)', output)
    ap_lwapp_uptime = re.search(r'AP\sLW\S+\s\w+\s\S+\s(.*)', output)
    join_date_time = re.search(r'Join\sDate\s\w+\s\S+\s(.*)', output)
    join_taken_time = re.search(r'Join\sTa\w+\s\S+\s(.*)', output)
    # Radio Section #
    five_power_config = re.search(r'Tx\sPower\sC\S+\s\S+\s(.*)', five_radio)
    five_tx_level = re.search(r'Current\sTx\s\S+\s\S+\s\S+\s(.*)', five_radio)
    five_channel = re.search(r'Current\sCh\S+\s\S+\s(.*)', five_radio)
    five_channel_width = re.search(r'Channel\sW\S+\s(.*)', five_radio)
    #
    two_power_config = re.search(r'Tx\sPower\sC\S+\s\S+\s(.*)', two_radio)
    two_tx_level = re.search(r'Current\sTx\s\S+\s\S+\s\S+\s(.*)', two_radio)
    two_channel = re.search(r'Current\sCh\S\S+\s\S+\s(.*)', two_radio)
    two_channel_width = re.search(r'Channel\sW\S+\s(.*)', two_radio)

    print('=> The AP name is =', ap_name.group(1))
    print('=> The AP mac-add =', ap_mac.group(1))
    print('=> The IPv4 address config =', ipv4_mode.group(1))
    print('=> The AP IPv4 IP =', ipv4.group(1))
    print('=> The AP netmask is =', netmask.group(1))
    print('=> The AP gateway =', gw.group(1))
    print('=> The AP CAPWAP path MTU =', capwap_path_mtu.group(1))
    print('=> The AP location =', location.group(1))
    print('=> The AP group name =', group_name.group(1))
    print('=> The AP Admin state =', admin_state.group(1))
    print('=> The AP Operation mode =', operation_mode.group(1))
    print('=> The AP mode =', ap_mode.group(1))
    print('=> The AP software version =', soft_ver.group(1))
    print('=> The AP Model =', ap_model.group(1))
    print('=> The AP image =', ap_image.group(1))
    print('=> The AP S/N =', sn.group(1))
    print('=> The AP uptime =', ap_uptime.group(1))
    print('=> The AP LWAPP uptime is =', ap_lwapp_uptime.group(1))
    print('=> The Join Data and Time =', join_date_time.group(1))
    print('=> The Join Taken Time =', join_taken_time.group(1))
    print('*---*-*---*-*---*-*---*-*---*')
    print('==> 2.4Ghz info <==')
    print('=>AP connected to channel =', two_channel.group(1))
    print('=>AP 2.4Ghz channel width =', two_channel_width.group(1))
    print('=>AP 2.4Ghz Tx-Power-Level =', two_tx_level.group(1))
    print('=>AP 2.4Ghz Tx-Power-Level-Assignment=', two_power_config.group(1))
    print('==> 5Ghz info <==')
    print('=>AP connected to channel =', five_channel.group(1))
    print('=>AP 5Ghz channel width =', five_channel_width.group(1))
    print('=>AP 5Ghz Tx-Power-Level =', five_tx_level.group(1))
    print('=>AP 5Ghz Tx-Power-Level-Assignment=', five_power_config.group(1))


def get_ping_info_rates(ping, ip, net_connect):
    if '!' in ping:
        output = re.search(r'rate\s\S+\s(\S+)\s\S+\s.(\d.\d)..\s\S+\s\S+\s.\s(\d+).(\d+).(\d+)', ping)
        print('=> Date =', get_time_date()[0], '=> Time =', get_time_date()[1])
        print('==> Ping results to ' + ip)
        print('=> Success rate =', output.group(1), '% ', output.group(2))
        print('=> round-trip min time', output.group(3), 'msec')
        print('=> round-trip avg time', output.group(4), 'msec')
        print('=> round-trip max time', output.group(5), 'msec')
    else:
        print('=> ICMP ping failed')


def wlc_clients_associated(ap_name, net_connect):
    # date and time
    print('==> Date =', get_time_date()[0], '=> Time =', get_time_date()[1])
    # getting output for 5Ghz
    print("==> Getting 5Ghz clients associates with", ap_name)
    output = net_connect.send_command("show client ap 802.11a" + " " + ap_name)
    # filtering output for 5Ghz
    five = re.findall(r'[a-fA-F0-9]{2}:.*', output)
    print('MAC Address        AP Id   Status         WLAN Id    Authenticated')
    for j in five:
        print(j)

    print('=> Total 5GHz clients connected =', len(five))
    print('\n')

    # getting output for 2.4Ghz
    print("==> Getting 2.4Ghz clients associated with", ap_name)
    output1 = net_connect.send_command("show client ap 802.11b" + " " + ap_name)
    print('MAC Address        AP Id   Status         WLAN Id    Authenticated')
    # filtering output for 2.4Ghz
    two = re.findall(r'[a-fA-F0-9]{2}:.*', output1)
    for c in two:
        print(c)
    print('=> Total 2.4GHz clients connected =', len(two))
    print('==> Total clients connected to ' + ap_name + " " + '=', len(five) + len(two))


def get_ios_wlc_ap(apname, net_connect):
    output = net_connect.send_command("sh ap name " + apname + " " + 'config general', read_timeout=603)
    if 'Invalid AP Name' in output:
        print('==> AP ' + apname + " " + "isn't joined to WLC")
    else:
        x = ' '
        lthree_info = re.search(
            r'802.11a\s+.\s.*[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)',
            output)
        ap_uptime = re.search(r'Cisco\sAP\sSec.*[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)',
                              output)
        ap_name = re.search(r'AP\sN\w+\s+.\s(.*)', output).group(1)
        # name
        print('Cisco AP Name', x * 33, ':', ap_name)
        # country code
        print(re.search(r'AP\sCou\S.*', output).group())
        # mac-add
        print(lthree_info.group(1))
        # ip address config
        print(lthree_info.group(2))
        # IP address
        print(lthree_info.group(3))
        # netmask
        print(lthree_info.group(4))
        # gateway
        print(lthree_info.group(5))
        # capwap mtu tunnel
        print(re.search(r'CAP\w+\sPa.*', output).group())
        # ssh state
        print(re.search(r'SSH.*', output).group())
        # ap location
        print(re.search(r'Cisco\sAP\sLo.*', output).group())
        # site tag name
        print(re.search(r'Site.*', output).group())
        # RF tag name
        print(re.search(r'RF\sT.*', output).group())
        # policy tag name
        print(re.search(r'Policy.*', output).group())
        # ap join profile
        print(re.search(r'AP\sjoi.*', output).group())
        # flex profile
        print(re.search(r'Fle.*', output).group())
        # admin state
        print(re.search(r'Admini.*', output).group())
        #
        print(re.search(r'Opera.*', output).group())
        # AP mode
        print(re.search(r'AP\sMo.*', output).group())
        # ap vlan tagging
        print(re.search(r'AP\sVL\w+\stagg.*', output).group())
        # AP vlan tag
        print(re.search(r'AP\sVL\w+\stag\s.*', output).group())
        # soft version
        print(re.search(r'So.*', output).group())
        # boot version
        print(re.search(r'Boo.*', output).group())
        # LED state
        print(re.search(r'LED\sS.*', output).group())
        # AP model
        print(re.search(r'.*Model.*', output).group())
        # Ap serial #
        print(re.search(r'.*Seri.*', output).group())
        # AP username
        print(re.search(r'AP\sUs.*', output).group())
        # AP 802.1x user mode
        print(re.search(r'AP\s802\S+\sU\w+\sM.*', output).group())
        # AP uptime
        print(ap_uptime.group(1))
        # AP capwap uptime
        print(ap_uptime.group(2))
        # AP join date and time
        print(ap_uptime.group(3))
        # AP join taken time
        print(ap_uptime.group(4))
        # tcp mss adjust
        print(re.search(r'AP\sT\w+\sM\w+\sA.*', output).group())
        # tcp mss size
        print(re.search(r'AP\sT\w+\sM\w+\sS.*', output).group())
        #


def get_wlc_facts(net_connect):
    output = net_connect.send_command("show ver")
    if 'Incorrect usage' in output:
        sysinfo = net_connect.send_command("show sysinfo")
        inv = net_connect.send_command("show inventory")
        model = re.search(r'PID.\s(\w+.\w+.\w+)', inv)
        name = re.search(r'System\sNam\w+\S+\s(.*)', sysinfo)
        os_ver = re.search(r'Prod\w+\sVer\w+\S+\s(.*)', sysinfo)
        uptime = re.search(r'Up\sT\w+\S+\s(.*)', sysinfo)
        wlans_total = re.search(r'WLA\w+\S+\s(\d+)', sysinfo)
        total_clients = re.search(r'Clien\w+\S+\s(\d+)', sysinfo)
        print('*---*-*---*-*---*-*---*-*---*')
        print('===> Date =', get_time_date()[0], '=> Time =', get_time_date()[1])
        print('==> Platform AireOS =>', model.group(1))
        print('=> System Name:', name.group(1))
        print('=> OS version:', os_ver.group(1))
        print('=> UPTIME:', uptime.group(1))
        print('=> Total of WLANs:', wlans_total.group(1))
        print('=> Total Clients:', total_clients.group(1))

    elif 'Cisco IOS XE':
        inv_9800 = net_connect.send_command("show inventory")
        model = re.search(r'Cha\w+\s\d+".*[\r\n]+([^\r\n]+)', inv_9800)
        model_inv = re.search(r'P\w+.\s(\S+)', model.group(1))
        name = re.search(r'(.*)\s\buptime\b', output)
        os = re.search(r'Ci\w+\sIOS\sXE\s\S+\s\S+\s(.*)', output)
        uptime = re.search(r'uptime\s\S+\s(.*)', output)
        serial1 = re.search(r'Chassis\s1".*[\r\n]+([^\r\n]+)', inv_9800)
        ser = re.search(r'SN.\s(\S+)', serial1.group())
        serial2 = re.search(r'Chassis\s2".*[\r\n]+([^\r\n]+)', inv_9800)
        ser2 = re.search(r'SN.\s(\S+)', serial2.group())
        print('*---*-*---*-*---*-*---*-*---*')
        print('===> Date =', get_time_date()[0], '=> Time =', get_time_date()[1])
        print('==> Platform IOS =>', 'Model=' + model_inv.group(1))
        print('=> System Name:', name.group(1))
        print('=> OS version:', os.group(1))
        print('=> UPTIME:', uptime.group(1))
        if ser2.group(1):
            print('=> Serial #(s) =', '[' + ser.group(1), ',' + ser2.group(1) + ']')
        else:
            print('=> Serial # =', ser.group(1))


def get_wlc_wlan_qos(net_connect):
    # getting MedNet or CrewNet ID
    ssid_id = net_connect.send_command("show wlan summa", read_timeout=603)
    MedNet = re.search(r'(\d+).*\bMedallionNet\b', ssid_id)
    CreNet = re.search(r'(\d+).*\bCrewNet\b', ssid_id)
    print('=> MedNet ID =', MedNet.group(1))
    print('=> CrewNet ID =', CreNet.group(1))
    print('*---*-*---*-*---*-*---*-*---*')

    # getting wlan info MedNet
    print('==> Medallion Info <==')
    wlan_MedNet = net_connect.send_command("show wlan" + " " + MedNet.group(1), read_timeout=603)
    ssid = re.search(r'Name\s\(SSID\)\S+\s(\w+)', wlan_MedNet)
    ssid_status = re.search(r'Status\S+\s(\w+)', wlan_MedNet)
    broadcast = re.search(r'Broad\w+\sS\w+\S+\s(\w+)', wlan_MedNet)
    qos = re.search(r'Qual\w+\s\w+\s\w+\S+\s(\w+)', wlan_MedNet)
    total_clients = re.search(r'Acti\w+\sC\w+\S+\s(\d+)', wlan_MedNet)

    print('=> SSID =', ssid.group(1))
    print('=> Active Clients =', total_clients.group(1))
    print('=> SSID Status =', ssid_status.group(1))
    print('=> Broadcast SSID =', broadcast.group(1))
    print('=> Quality of Service =', qos.group(1))
    print('==> Per-SSID Rate Limits     Upstream             Downstream')
    x = ' '
    # per SSID
    per_wlan = re.search(r'Per-SSID Rate Limits.*[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)',
                         wlan_MedNet)
    wlc_ave_data_rate = re.search(r'Ave\w+\sD\w+\s\w+\S+\s+(\d+)\s+(\d+)', per_wlan.group(1))
    wlc_real_time_data_rate = re.search(r'Ave\w+\sRe\w+\sD\w+\sR\w+\S+\s+(\d+)\s+(\d+)', per_wlan.group(2))
    wlc_burst_data_rate = re.search(r'Bur\w+\sDa\w+\sR\w+\S+\s+(\d+)\s+(\d+)', per_wlan.group(3))
    wlc_burst_realtime_data_rate = re.search(r'Bur\w+\sRe\w+\sD\w+\sRa\w+\S+\s+(\d+)\s+(\d+)', per_wlan.group(4))
    print('=> Average Data Rate           ', wlc_ave_data_rate.group(1) + ' kbps', x * 12,
          wlc_ave_data_rate.group(2) + ' kbps')
    print('=> Average Realtime Data Rate  ', wlc_real_time_data_rate.group(1) + ' kbps', x * 12,
          wlc_real_time_data_rate.group(2) + ' kbps')
    print('=> Burst Data Rate             ', wlc_burst_data_rate.group(1) + ' kbps', x * 12,
          wlc_burst_data_rate.group(2) + ' kbps')
    print('=> Burst Realtime Data Rate    ', wlc_burst_realtime_data_rate.group(1) + ' kbps', x * 12,
          wlc_burst_realtime_data_rate.group(2) + ' kbps')

    print('==> Per-Client Rate Limits     Upstream             Downstream')
    # per Clients
    per_clients = re.search(
        r'Per-Client Rate Limits.*[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)', wlan_MedNet)
    mednet_wlc_ave_data_rate = re.search(r'Ave\w+\sD\w+\s\w+\S+\s+(\d+)\s+(\d+)', per_clients.group(1))
    mednet_wlc_real_time_data_rate = re.search(r'Ave\w+\sRe\w+\sD\w+\sR\w+\S+\s+(\d+)\s+(\d+)', per_clients.group(2))
    mednet_wlc_burst_data_rate = re.search(r'Bur\w+\sDa\w+\sR\w+\S+\s+(\d+)\s+(\d+)', per_clients.group(3))
    mednet_wlc_burst_realtime_data_rate = re.search(r'Bur\w+\sRe\w+\sD\w+\sRa\w+\S+\s+(\d+)\s+(\d+)',
                                                    per_clients.group(4))

    print('=> Average Data Rate           ', mednet_wlc_ave_data_rate.group(1) + ' kbps', x * 12,
          mednet_wlc_ave_data_rate.group(2) + ' kbps')
    print('=> Average Realtime Data Rate  ', mednet_wlc_real_time_data_rate.group(1) + ' kbps', x * 12,
          mednet_wlc_real_time_data_rate.group(2) + ' kbps')
    print('=> Burst Data Rate             ', mednet_wlc_burst_data_rate.group(1) + ' kbps', x * 12,
          mednet_wlc_burst_data_rate.group(2) + ' kbps')
    print('=> Burst Realtime Data Rate    ', mednet_wlc_burst_realtime_data_rate.group(1) + ' kbps', x * 12,
          mednet_wlc_burst_realtime_data_rate.group(2) + ' kbps')

    # getting wlan info CrewNet
    print('*---*-*---*-*---*-*---*-*---*')
    print('==> CrewNet Info <==')
    wlan_CrewNet = net_connect.send_command("show wlan" + " " + CreNet.group(1), read_timeout=603)
    crewnet_ssid = re.search(r'Name\s\(SSID\)\S+\s(\w+)', wlan_CrewNet)
    crewnet_ssid_status = re.search(r'Status\S+\s(\w+)', wlan_CrewNet)
    crewnet_broadcast = re.search(r'Broad\w+\sS\w+\S+\s(\w+)', wlan_CrewNet)
    crewnet_qos = re.search(r'Qual\w+\s\w+\s\w+\S+\s(\w+)', wlan_CrewNet)
    crewnet_total_clients = re.search(r'Acti\w+\sC\w+\S+\s(\d+)', wlan_CrewNet)

    print('=> SSID =', crewnet_ssid.group(1))
    print('=> Active Clients =', crewnet_total_clients.group(1))
    print('=> SSID Status =', crewnet_ssid_status.group(1))
    print('=> Broadcast SSID =', crewnet_broadcast.group(1))
    print('=> Quality of Service =', crewnet_qos.group(1))
    print('==> Per-SSID Rate Limits     Upstream             Downstream')

    per_wlan_ssid = re.search(
        r'Per-SSID Rate Limits.*[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)',
        wlan_CrewNet)
    wlc_ssid_ave_data_rate = re.search(r'Ave\w+\sD\w+\s\w+\S+\s+(\d+)\s+(\d+)', per_wlan_ssid.group(1))
    wlc_ssid_real_time_data_rate = re.search(r'Ave\w+\sRe\w+\sD\w+\sR\w+\S+\s+(\d+)\s+(\d+)', per_wlan_ssid.group(2))
    wlc_ssid_burst_data_rate = re.search(r'Bur\w+\sDa\w+\sR\w+\S+\s+(\d+)\s+(\d+)', per_wlan_ssid.group(3))
    wlc_ssid_burst_realtime_data_rate = re.search(r'Bur\w+\sRe\w+\sD\w+\sRa\w+\S+\s+(\d+)\s+(\d+)',
                                                  per_wlan_ssid.group(4))
    print('=> Average Data Rate           ', wlc_ssid_ave_data_rate.group(1) + ' kbps', x * 12,
          wlc_ssid_ave_data_rate.group(2) + ' kbps')
    print('=> Average Realtime Data Rate  ', wlc_ssid_real_time_data_rate.group(1) + ' kbps', x * 12,
          wlc_ssid_real_time_data_rate.group(2) + ' kbps')
    print('=> Burst Data Rate             ', wlc_ssid_burst_data_rate.group(1) + ' kbps', x * 12,
          wlc_ssid_burst_data_rate.group(2) + ' kbps')
    print('=> Burst Realtime Data Rate    ', wlc_ssid_burst_realtime_data_rate.group(1) + ' kbps', x * 12,
          wlc_ssid_burst_realtime_data_rate.group(2) + ' kbps')

    # per Clients
    print('==> Per-Client Rate Limits      Upstream              Downstream')
    per_wlan = re.search(
        r'Per-Client Rate Limits.*[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)',
        wlan_CrewNet)
    crewnet_wlc_ave_data_rate = re.search(r'Ave\w+\sD\w+\s\w+\S+\s+(\d+)\s+(\d+)', per_wlan.group(1))
    crewnet_wlc_real_time_data_rate = re.search(r'Ave\w+\sRe\w+\sD\w+\sR\w+\S+\s+(\d+)\s+(\d+)', per_wlan.group(2))
    crewnet_wlc_burst_data_rate = re.search(r'Bur\w+\sDa\w+\sR\w+\S+\s+(\d+)\s+(\d+)', per_wlan.group(3))
    crewnet_wlc_burst_realtime_data_rate = re.search(r'Bur\w+\sRe\w+\sD\w+\sRa\w+\S+\s+(\d+)\s+(\d+)',
                                                     per_wlan.group(4))

    print('=> Average Data Rate           ', crewnet_wlc_ave_data_rate.group(1) + ' kbps', x * 12,
          crewnet_wlc_ave_data_rate.group(2) + ' kbps')
    print('=> Average Realtime Data Rate  ', crewnet_wlc_real_time_data_rate.group(1) + ' kbps', x * 12,
          crewnet_wlc_real_time_data_rate.group(2) + ' kbps')
    print('=> Burst Data Rate             ', crewnet_wlc_burst_data_rate.group(1) + ' kbps', x * 12,
          crewnet_wlc_burst_data_rate.group(2) + ' kbps')
    print('=> Burst Realtime Data Rate    ', crewnet_wlc_burst_realtime_data_rate.group(1) + ' kbps', x * 12,
          crewnet_wlc_burst_realtime_data_rate.group(2) + ' kbps' + '\n')

    set_qos = input(
        "==> Do you want to change QoS values on MedNet and CrewNet, (Y) to Continue, (N) to Cancel: ").lower()
    print('\n')
    if set_qos in yes_option:
        print('*****======> Changing QoS values for MedNet <======*****')
        set_wlc_qos(MedNet.group(1), net_connect)
        print('\n')

        # Calling for CrewNet
        print('*****======> Changing QoS values for CrewNet <======*****')
        set_wlc_qos(CreNet.group(1), net_connect)

    elif set_qos in no_option:
        print("==> No Qos Changes applied <==")


def get_wlc_wlan_qos_crewcompass(net_connect):
    # checking WLC family
    wlc_family = net_connect.send_command("show version")
    if 'Incorrect usage' in wlc_family:
        print('==> Using aireOS WLC <==')
        ssid_id = net_connect.send_command("show wlan summa", read_timeout=603)
        CrewCompass = re.search(r'(\d+).*\bCrewCompass\b', ssid_id)
        print('=> CrewCompass ID =', CrewCompass.group(1))
        print('*---*-*---*-*---*-*---*-*---*')
        # getting wlan info MedNet
        print('==> CrewCompass Info <==')
        wlan_CrewCompass = net_connect.send_command("show wlan" + " " + CrewCompass.group(1), read_timeout=603)
        ssid = re.search(r'Name\s\(SSID\)\S+\s(\w+)', wlan_CrewCompass)
        ssid_status = re.search(r'Status\S+\s(\w+)', wlan_CrewCompass)
        broadcast = re.search(r'Broad\w+\sS\w+\S+\s(\w+)', wlan_CrewCompass)
        qos = re.search(r'Qual\w+\s\w+\s\w+\S+\s(\w+)', wlan_CrewCompass)
        total_clients = re.search(r'Acti\w+\sC\w+\S+\s(\d+)', wlan_CrewCompass)
        print('=> SSID =', ssid.group(1))
        print('=> Active Clients =', total_clients.group(1))
        print('=> SSID Status =', ssid_status.group(1))
        print('=> Broadcast SSID =', broadcast.group(1))
        print('=> Quality of Service =', qos.group(1))
        print('==> Per-SSID Rate Limits     Upstream             Downstream')
        x = ' '
        # per SSID
        per_wlan = re.search(
            r'Per-SSID Rate Limits.*[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)',
            wlan_CrewCompass)
        wlc_ave_data_rate = re.search(r'Ave\w+\sD\w+\s\w+\S+\s+(\d+)\s+(\d+)', per_wlan.group(1))
        wlc_real_time_data_rate = re.search(r'Ave\w+\sRe\w+\sD\w+\sR\w+\S+\s+(\d+)\s+(\d+)', per_wlan.group(2))
        wlc_burst_data_rate = re.search(r'Bur\w+\sDa\w+\sR\w+\S+\s+(\d+)\s+(\d+)', per_wlan.group(3))
        wlc_burst_realtime_data_rate = re.search(r'Bur\w+\sRe\w+\sD\w+\sRa\w+\S+\s+(\d+)\s+(\d+)', per_wlan.group(4))
        print('=> Average Data Rate           ', wlc_ave_data_rate.group(1) + ' kbps', x * 12,
              wlc_ave_data_rate.group(2) + ' kbps')
        print('=> Average Realtime Data Rate  ', wlc_real_time_data_rate.group(1) + ' kbps', x * 12,
              wlc_real_time_data_rate.group(2) + ' kbps')
        print('=> Burst Data Rate             ', wlc_burst_data_rate.group(1) + ' kbps', x * 12,
              wlc_burst_data_rate.group(2) + ' kbps')
        print('=> Burst Realtime Data Rate    ', wlc_burst_realtime_data_rate.group(1) + ' kbps', x * 12,
              wlc_burst_realtime_data_rate.group(2) + ' kbps')

        print('==> Per-Client Rate Limits     Upstream             Downstream')
        # per Clients
        per_clients = re.search(
            r'Per-Client Rate Limits.*[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)',
            wlan_CrewCompass)
        mednet_wlc_ave_data_rate = re.search(r'Ave\w+\sD\w+\s\w+\S+\s+(\d+)\s+(\d+)', per_clients.group(1))
        mednet_wlc_real_time_data_rate = re.search(r'Ave\w+\sRe\w+\sD\w+\sR\w+\S+\s+(\d+)\s+(\d+)',
                                                   per_clients.group(2))
        mednet_wlc_burst_data_rate = re.search(r'Bur\w+\sDa\w+\sR\w+\S+\s+(\d+)\s+(\d+)', per_clients.group(3))
        mednet_wlc_burst_realtime_data_rate = re.search(r'Bur\w+\sRe\w+\sD\w+\sRa\w+\S+\s+(\d+)\s+(\d+)',
                                                        per_clients.group(4))

        print('=> Average Data Rate           ', mednet_wlc_ave_data_rate.group(1) + ' kbps', x * 12,
              mednet_wlc_ave_data_rate.group(2) + ' kbps')
        print('=> Average Realtime Data Rate  ', mednet_wlc_real_time_data_rate.group(1) + ' kbps', x * 12,
              mednet_wlc_real_time_data_rate.group(2) + ' kbps')
        print('=> Burst Data Rate             ', mednet_wlc_burst_data_rate.group(1) + ' kbps', x * 12,
              mednet_wlc_burst_data_rate.group(2) + ' kbps')
        print('=> Burst Realtime Data Rate    ', mednet_wlc_burst_realtime_data_rate.group(1) + ' kbps', x * 12,
              mednet_wlc_burst_realtime_data_rate.group(2) + ' kbps')
        print('*---*-*---*-*---*-*---*-*---*')

    elif 'Cisco IOS XE Software' in wlc_family:
        print('==> Using IOS (9800) WLC <==')
        # getting CrewCompass ID from wlc
        ssid_id = net_connect.send_command("show wlan summa")
        crewcompass = re.search(r'(\d+).*\bCrewCompass\b', ssid_id)
        print('=> CrewCompass ID =', crewcompass.group(1))
        print('*---*-*---*-*---*-*---*-*---*')
        # getting wlan info CrewCompass
        wlan_CrewCompass = net_connect.send_command("show wlan id" + " " + crewcompass.group(1))
        crewcompass_profile = net_connect.send_command('sho wireless profile policy detailed Crewcompass_profile')
        ssid = re.search(r'Name\s\(SSID\)\s+.\s(\w+)', wlan_CrewCompass)
        ssid_status = re.search(r'Status\s+.\s(\w+)', wlan_CrewCompass)
        broadcast = re.search(r'Broad\w+\s\w+\s+.\s(\w+)', wlan_CrewCompass)
        total_clients = re.search(r'Acti\w+\sCl\w+\s+.\s(\d+)', wlan_CrewCompass)
        auth = re.search(r'802.11\sAu\w+\s+.\s(.*)', wlan_CrewCompass)

        print('==> CrewCompass Info <==')
        print('=> SSID =', ssid.group(1))
        print('=> SSID Status =', ssid_status.group(1))
        print('=> Broadcast SSID =', broadcast.group(1))
        print('=> Total Clients connected', total_clients.group(1))
        print('=> 802.11 Authentication =', auth.group(1))
        print('==> QOS per SSID')
        crewcompass_qos_ssid = re.search('QOS\sper\sS.*[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)', crewcompass_profile)
        ingres_crewcompass = re.search('In\S+\s\S+\s\S+\s+.\s(.*)', crewcompass_qos_ssid.group(1))
        egress_crewcompass = re.search('Eg\S+\s\S+\s\S+\s+.\s(.*)', crewcompass_qos_ssid.group(2))
        if 'Not' in ingres_crewcompass.group(1):
            print('=> Ingress = Not Configured')
        else:
            print('=> Ingress =', ingres_crewcompass.group(1))

        if 'Not' in egress_crewcompass.group(1):
            print('=> Egress = Not Configured')
        else:
            print('=> Egress  =', egress_crewcompass.group(1))

        print('==> QOS per Client')
        crewcompass_qos_client = re.search('QOS\sper\sC.*[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)', crewcompass_profile)
        ingres_client = re.search('In\S+\s\S+\s\S+\s+.\s(.*)', crewcompass_qos_client.group(1))
        egress_client = re.search('Eg\S+\s\S+\s\S+\s+.\s(.*)', crewcompass_qos_client.group(2))
        if 'Not' in ingres_client.group(1):
            print('=> Ingress = Not Configured')
        else:
            print('=> Ingress =', ingres_client.group(1))

        if 'Not' in egress_client.group(1):
            print('=> Egress = Not Configured')
        else:
            print('=> Egress  =', egress_client.group(1))
        print('*---*-*---*-*---*-*---*-*---*')


def get_wlc_wlan_qos_9800(net_connect):
    # getting MedNet or CrewNet ID
    ssid_id = net_connect.send_command("show wlan summa")
    MedNet = re.search(r'(\d+).*\bMedallionNet\b', ssid_id)
    CreNet = re.search(r'(\d+).*\bCrewNet\b', ssid_id)
    print('=> MedNet ID =', MedNet.group(1))
    print('=> CrewNet ID =', CreNet.group(1))
    print('*---*-*---*-*---*-*---*-*---*')

    # getting wlan info MedNet
    wlan_MedNet = net_connect.send_command("show wlan id" + " " + MedNet.group(1))
    MedNet_profile = net_connect.send_command('sho wireless profile policy detailed MedallionNet_Policy')
    ssid = re.search(r'Name\s\(SSID\)\s+.\s(\w+)', wlan_MedNet)
    ssid_status = re.search(r'Status\s+.\s(\w+)', wlan_MedNet)
    broadcast = re.search(r'Broad\w+\s\w+\s+.\s(\w+)', wlan_MedNet)
    total_clients = re.search(r'Acti\w+\sCl\w+\s+.\s(\d+)', wlan_MedNet)
    auth = re.search(r'802.11\sAu\w+\s+.\s(.*)', wlan_MedNet)

    print('==> MedallionNet Info <==')
    print('=> SSID =', ssid.group(1))
    print('=> SSID Status =', ssid_status.group(1))
    print('=> Broadcast SSID =', broadcast.group(1))
    print('=> Total Clients connected', total_clients.group(1))
    print('=> 802.11 Authentication =', auth.group(1))
    print('==> QOS per SSID')
    mednet_qos_ssid = re.search('QOS\sper\sS.*[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)', MedNet_profile)
    ingres_mednet = re.search('In\S+\s\S+\s\S+\s+.\s(.*)', mednet_qos_ssid.group(1))
    egress_mednet = re.search('Eg\S+\s\S+\s\S+\s+.\s(.*)', mednet_qos_ssid.group(2))
    if 'Not' in ingres_mednet.group(1):
        print('=> Ingress = Not Configured')
    else:
        print('=> Ingress =', ingres_mednet.group(1))

    if 'Not' in egress_mednet.group(1):
        print('=> Egress = Not Configured')
    else:
        print('=> Egress  =', egress_mednet.group(1))

    print('==> QOS per Client')
    mednet_qos_client = re.search('QOS\sper\sC.*[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)', MedNet_profile)
    ingres_client = re.search('In\S+\s\S+\s\S+\s+.\s(.*)', mednet_qos_client.group(1))
    egress_client = re.search('Eg\S+\s\S+\s\S+\s+.\s(.*)', mednet_qos_client.group(2))
    if 'Not' in ingres_client.group(1):
        print('=> Ingress = Not Configured')
    else:
        print('=> Ingress =', ingres_client.group(1))

    if 'Not' in egress_client.group(1):
        print('=> Egress = Not Configured')
    else:
        print('=> Egress  =', egress_client.group(1))

    # getting wlan info CrewNet
    print('*---*-*---*-*---*-*---*-*---*')
    wlan_CrewNet = net_connect.send_command("show wlan id" + " " + CreNet.group(1))
    CrewNet_profile = net_connect.send_command('sho wireless profile policy detailed MedallionCrew_Profile')
    ssid = re.search(r'Name\s\(SSID\)\s+.\s(\w+)', wlan_CrewNet)
    ssid_status = re.search(r'Status\s+.\s(\w+)', wlan_CrewNet)
    broadcast = re.search(r'Broad\w+\s\w+\s+.\s(\w+)', wlan_CrewNet)
    total_clients = re.search(r'Acti\w+\sCl\w+\s+.\s(\d+)', wlan_CrewNet)
    auth = re.search(r'802.11\sAu\w+\s+.\s(.*)', wlan_CrewNet)

    print('==> CrewNet Info <==')
    print('=> SSID =', ssid.group(1))
    print('=> SSID Status =', ssid_status.group(1))
    print('=> Broadcast SSID =', broadcast.group(1))
    print('=> Total Clients connected', total_clients.group(1))
    print('=> 802.11 Authentication =', auth.group(1))
    print('==> QOS per SSID')
    crewnet_qos_ssid = re.search('QOS\sper\sS.*[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)', CrewNet_profile)
    ingres_crewnet = re.search('In\S+\s\S+\s\S+\s+.\s(.*)', crewnet_qos_ssid.group(1))
    egress_crewnet = re.search('Eg\S+\s\S+\s\S+\s+.\s(.*)', crewnet_qos_ssid.group(2))
    if 'Not' in ingres_crewnet.group(1):
        print('=> Ingress = Not Configured')
    else:
        print('=> Ingress =', ingres_crewnet.group(1))

    if 'Not' in egress_crewnet.group(1):
        print('=> Egress = Not Configured')
    else:
        print('=> Egress  =', egress_crewnet.group(1))

    print('==> QOS per Client')
    crewnet_qos_client = re.search('QOS\sper\sC.*[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)', CrewNet_profile)
    ingres_client_crew = re.search('In\S+\s\S+\s\S+\s+.\s(.*)', crewnet_qos_client.group(1))
    egress_client_crew = re.search('Eg\S+\s\S+\s\S+\s+.\s(.*)', crewnet_qos_client.group(2))
    if 'Not' in ingres_client_crew.group(1):
        print('=> Ingress = Not Configured')
    else:
        print('=> Ingress =', ingres_client_crew.group(1))

    if 'Not' in egress_client_crew.group(1):
        print('=> Egress = Not Configured')
    else:
        print('=> Egress  =', egress_client_crew.group(1))


def wlc_get_api_cisco_prime_aireOS(ap_name):
    print("==> AP is not joined WLC")
    # checking with prime for details
    try:
        yp = re.match('^YP', ap_name)
        gp = re.match('^GP', ap_name)
        rp = re.match('^RP', ap_name)
        mj = re.match('^MJ', ap_name)
        ap = re.match('^AP', ap_name)
        cb = re.match('^CB', ap_name)
        co = re.match('^CO', ap_name)
        di = re.match('^DI', ap_name)
        ep = re.match('^EP', ap_name)
        kp = re.match('^KP', ap_name)
        ip = re.match('^IP', ap_name)
        ru = re.match('^RU', ap_name)
        sa = re.match('^SA', ap_name)
        pev2 = re.match('^PEV', ap_name)

    except AttributeError:
        pass

    try:
        if yp.group():
            pi_ip = '10.125.36.196'
            print('=> Trying to get info from Sky Prime')
            # connecting PI
            cisco_prime_api_results(ap_name, pi_ip)
            exit(0)

    except AttributeError:
        pass

    try:
        if gp.group():
            pi_ip = '10.122.227.196'
            print('=> Trying to get info from Regal Prime')
            # connecting PI
            cisco_prime_api_results(ap_name, pi_ip)
            exit(0)

    except AttributeError:
        pass

    try:
        if rp.group():
            pi_ip = '10.123.36.196'
            print('=> Trying to get info from Royal Prime')
            # connecting PI
            cisco_prime_api_results(ap_name, pi_ip)
            exit(0)

    except AttributeError:
        pass

    try:
        if mj.group():
            pi_ip = '10.124.164.196'
            print('=> Trying to get info from Majestic Prime')
            # connecting PI
            cisco_prime_api_results(ap_name, pi_ip)
            exit(0)

    except AttributeError:
        pass

    try:
        if ap.group():
            pi_ip = '10.121.228.196'
            print('=> Trying to get info from Grand Prime')
            # connecting PI
            cisco_prime_api_results(ap_name, pi_ip)
            exit(0)

    except AttributeError:
        pass

    try:
        if cb.group():
            pi_ip = '10.120.36.196'
            print('=> Trying to get info from Caribbean Prime')
            # connecting PI
            cisco_prime_api_results(ap_name, pi_ip)
            exit(0)

    except AttributeError:
        pass

    try:
        if co.group():
            pi_ip = '10.120.100.196'
            print('=> Trying to get info from Coral Prime')
            # connecting PI
            cisco_prime_api_results(ap_name, pi_ip)
            exit(0)

    except AttributeError:
        pass

    try:
        if di.group():
            pi_ip = '10.121.36.196'
            print('=> Trying to get info from Diamond Prime')
            # connecting PI
            cisco_prime_api_results(ap_name, pi_ip)
            exit(0)

    except AttributeError:
        pass

    try:
        if ep.group():
            pi_ip = '10.121.100.196'
            print('=> Trying to get info from Emerald Prime')
            # connecting PI
            cisco_prime_api_results(ap_name, pi_ip)
            exit(0)

    except AttributeError:
        pass

    try:
        if kp.group():
            pi_ip = '10.120.164.196'
            print('=> Trying to get info from Crown Prime at ' + pi_ip)
            # connecting PI
            cisco_prime_api_results(ap_name, pi_ip)
            exit(0)

    except AttributeError:
        pass

    try:
        if ip.group():
            pi_ip = '10.122.36.196'
            print('=> Trying to get info from Island Prime at ' + pi_ip)
            # connecting PI
            cisco_prime_api_results(ap_name, pi_ip)
            exit(0)

    except AttributeError:
        pass

    try:
        if ru.group():
            pi_ip = '10.123.100.196'
            print('=> Trying to get info from Ruby Prime at ' + pi_ip)
            # connecting PI
            cisco_prime_api_results(ap_name, pi_ip)
            exit(0)

    except AttributeError:
        pass

    try:
        if sa.group():
            pi_ip = '10.123.164.196'
            print('=> Trying to get info from Sapphire Prime at ' + pi_ip)
            # connecting PI
            cisco_prime_api_results(ap_name, pi_ip)
            exit(0)

    except AttributeError:
        pass

    try:
        if pev2.group():
            print('=> There is no Prime for PEv2')
            exit(0)

    except AttributeError:
        pass


def wlc_get_api_cisco_prime_ios(ap_name):
    print("==> AP is not joined WLC")
    # checking with prime for details
    try:
        ex = re.match('^EX', ap_name)
        xp = re.match('^XP', ap_name)
        xic = re.match('^XIC', ap_name)
        xic = re.match('^XIC2.5', ap_name)

    except AttributeError:
        pass

    try:
        if ex.group():
            pi_ip = '10.125.100.196'
            print('=> Trying to get info from Enchanted Prime at ' + pi_ip)
            # connecting PI
            cisco_prime_api_results(ap_name, pi_ip)
            exit(0)

    except AttributeError:
        pass

    try:
        if xp.group():
            pi_ip = '10.125.164.196'
            print('=> Trying to get info from Discovery Prime at ' + pi_ip)
            # connecting PI
            cisco_prime_api_results(ap_name, pi_ip)
            exit(0)

    except AttributeError:
        pass

    try:
        if xic.group():
            pi_ip = '10.126.100.196'
            print('=> Trying to get info from XIC Prime at ' + pi_ip)
            # connecting PI
            cisco_prime_api_results(ap_name, pi_ip)
            exit(0)

    except AttributeError:
        pass

    try:
        if xic.group():
            pi_ip = '10.126.100.196'
            print('=> Trying to get info from XIC2.5 Prime at ' + pi_ip)
            # connecting PI
            cisco_prime_api_results(ap_name, pi_ip)
            exit(0)

    except AttributeError:
        pass


def get_wlc_ap_facts(ap_name, net_connect):
    output = net_connect.send_command("show ap config general" + " " + ap_name, read_timeout=603)
    #    print(ch.send_command("show ap config general" + " " + ap_name))

    if "invalid" in output:
        wlc_get_api_cisco_prime_aireOS(ap_name)
        exit(0)

    else:
        print("==> AP is joined WLC, getting facts <==")

    # Filtering AP name
    ap = re.search(r'Cisco\sAP\sN.*', output)
    print(ap.group())

    # Filtering Country code
    c_code = re.search(r'AP\sCoun.*', output)
    print(c_code.group())

    # Filtering Regulatory Domain
    regulatory = re.search(r'AP\sReg.*', output)
    print(regulatory.group())

    # Filtering mac-add
    mac = re.search(r'MAC\sAddre.*', output)
    print(mac.group())

    # IP Address Configuration
    ip_dhcp = re.search(r'IP\sAdd\w+\s.*', output)
    print(ip_dhcp.group())

    # IP Address
    ip_add = re.search(r'IP\sAdd\w+(.){39}\s\d.*', output)
    print(ip_add.group())

    # IP netmask
    ip_net = re.search(r'IP\sNe.*', output)
    print(ip_net.group())

    # Gateway
    gateway = re.search(r'Gatew.*', output)
    print(gateway.group())

    # capwap
    capwap = re.search(r'CAPWAP\sP.*', output)
    print(capwap.group())

    # telnet
    telnet = re.search(r'Tel.*', output)
    print(telnet.group())

    # SSH
    ssh = re.search(r'Ssh.*', output)
    print(ssh.group())

    # ap location
    location = re.search(r'Ci\w+\sAP\sLo.*', output)
    print(location.group())

    # floor label
    floor = re.search(r'Ci\w+\sAP\sFl.*', output)
    print(floor.group())

    # group name
    group_name = re.search(r'Cis\w+\sAP\sG.*', output)
    print(group_name.group())

    # IP switch
    switch_pri = re.search(r'Primary\sC\w+\sS\w+\sI.*', output)
    print(switch_pri.group())

    # Tertiary
    tertiary = re.search(r'Tertiary\sCi\w+\sS\w+\sI.*', output)
    print(tertiary.group())

    # admin state
    admin_state = re.search(r'Administrative\s.*', output)
    print(admin_state.group())

    # mirror
    mirror = re.search(r'Mirror.*', output)
    print(mirror.group())

    # ap mode
    ap_mode = re.search(r'AP\sM\w+\s.*', output)
    print(ap_mode.group())

    # s/w version
    sw = re.search(r'S/W.*', output)
    print(sw.group())

    # boot
    boot = re.search(r'Boo.*', output)
    print(boot.group())

    # led state
    led = re.search(r'LED\sS.*', output)
    print(led.group())

    # power type
    power = re.search(r'Po\w+\sT.*', output)
    print(power.group())

    # ap model
    model = re.search(r'AP\sModel.*', output)
    print(model.group())

    # ap image
    ap_image = re.search(r'AP\sIma.*', output)
    print(ap_image.group())

    # ios version
    ios = re.search(r'IOS\sV\w+\S+.\d.*', output)
    print(ios.group())

    # ap serial #
    serial = re.search(r'AP\sSe.*', output)
    print(serial.group())

    # ap username
    ap_username = re.search(r'AP\sU\w+\sN.*', output)
    print(ap_username.group())

    # uptime
    uptime = re.search(r'AP\sUp\s.*', output)
    print(uptime.group())

    # ap LWAP
    ap_lwap = re.search(r'AP\sLW.*', output)
    print(ap_lwap.group())

    # join date
    join = re.search(r'Join\sD.*', output)
    print(join.group())

    # join taken
    join_taken = re.search(r'Join\sTa.*', output)
    print(join_taken.group())

    # memory type
    memory = re.search(r'Mem\w+\sT.*', output)
    print(memory.group())

    # memory size
    memory_size = re.search(r'Mem\w+\sS.*', output)
    print(memory_size.group())

    # flash size
    flash = re.search(r'Fla\w+\sS.*', output)
    print(flash.group())

    # ethernet duplex
    eth_duplex = re.search(r'Ether\w+\sP\w+\sD.*', output)
    print(eth_duplex.group())

    # ethernet speed
    eth_speed = re.search(r'Ether\w+\sP\w+\sS.*', output)
    print(eth_speed.group())

    # mss adjust
    mss_adjust = re.search(r'AP\sT\w+\sM\w+\sA.*', output)
    print(mss_adjust.group())

    # mss size
    mss_size = re.search(r'AP\sT\w+\sM\w+\sS.*', output)
    print(mss_size.group())

    # NSI
    nsi = re.search(r'NSI.*', output)
    print(nsi.group())

    # dot1x
    dot1 = re.search(r'AP\sDo\w+\sE.*', output)
    print(dot1.group())

    # testing email
    send_email(join.group())


def get_wlc_airos_clients_connected_by_ap(ap_name, net_connect):
    # getting output for 5Ghz
    five_tx = net_connect.send_command("show advanced 802.11a txpower")
    five_tx_filter = re.search(r'ap_name.*', five_tx)
    print(five_tx_filter)


def get_wlc_9800_clients_connected_by_ap(ap_name, net_connect):
    # getting output for 5Ghz
    print('*---*-*---*-*---*-*---*-*---*-*---*')
    # Getting tx_power for AP
    five_tx = net_connect.send_command("show ap dot11 5ghz summary | inc" + " " + ap_name, read_timeout=603)
    # five_details = re.search(r'^\w+\s+([\d|\w]+.[\d|\w]+.[\d|\w]+)\s+\d+\s+\w+\s+\w+\s+\d+\s+.(\d).\d\s(.\w+\s\w+.)\s+(.\d+.\d+.)', five_tx)
    five_details = re.search(
        r'(^NYC\w+|^EX\w+|^EX\S+|^XP\w+.\d+.\d+.\w+)\s+([\d|\w]+.[\d|\w]+.[\d|\w]+)\s+\d+\s+\w+\s+\w+\s+(\d+)\s+.(\d).\d\s(.\w+\s\w+.)\s+.(\d+).',
        five_tx)
    # printing details
    # getting date and time
    print('==> Date =', get_time_date()[0], '=> Time =', get_time_date()[1])
    if five_details.group(2):
        print('=> The AP mac-add is:', five_details.group(2))
    else:
        pass
    if five_details.group(4):
        print('=> The AP TX_POWER Level is:', five_details.group(4))
    else:
        pass
    if five_details.group(5):
        print('=> The AP TX_POWER dbm Level is:', five_details.group(5))
    else:
        pass
    if five_details.group(6):
        print('=> The AP 5Ghz Channel is:', five_details.group(6))
    else:
        pass
    if five_details.group(3):
        print('=> The 5Ghz Channel Width is:', five_details.group(3))
    else:
        pass
    print('*---*-*---*-*---*-*---*-*---*-*---*')

    print("==> Getting 5Ghz clients associates with", ap_name)
    output = net_connect.send_command("sh wireless client ap dot11 5ghz chassis active r0 | inc" + " " + ap_name,
                                      read_timeout=603)
    # filtering output for 5Ghz
    five = re.findall(r'.*', output)
    # five_new = [x.rstrip() for x in five] removing right trailing  spaces
    five_new = [ele for ele in five if ele.strip()]
    print('MAC Address        AP Id             Status          WLAN Id  Authenticated')
    for match in five_new:
        print(match)
    print('=> Total 5GHz clients connected =', len(five_new))

    print('\n')

    # getting output for 2.4Ghz
    print('*---*-*---*-*---*-*---*-*---*-*---*')
    # Getting tx_power for AP
    two_tx = net_connect.send_command("show ap dot11 24ghz summary | inc" + " " + ap_name, read_timeout=603)
    # two_details = re.search(r'^\w+\s+([\d|\w]+.[\d|\w]+.[\d|\w]+)\s+\d+\s+\w+\s+\w+\s+\d+\s+.(\d).\d\s(.\w+\s\w+.)\s+.(\d+).', two_tx)
    two_details = re.search(
        r'(^NYC\w+|^EX\w+|^EX\S+|^XP\w+.\d+.\d+.\w+)\s+([\d|\w]+.[\d|\w]+.[\d|\w]+)\s+\d+\s+\w+\s+\w+\s+(\d+)\s+.(\d).\d\s(.\w+\s\w+.)\s+.(\d+).',
        two_tx)
    # printing details
    # getting date and time
    print('==> Date =', get_time_date()[0], '=> Time =', get_time_date()[1])
    if two_details.group(2):
        print('=> The AP mac-add is:', two_details.group(2))
    else:
        pass
    if two_details.group(4):
        print('=> The AP TX_POWER Level is:', two_details.group(4))
    else:
        pass
    if two_details.group(5):
        print('=> The AP TX_POWER dbm Level is:', two_details.group(5))
    else:
        pass
    if two_details.group(6):
        print('=> The AP 2.4Ghz Channel is:', two_details.group(6))
    else:
        pass
    if two_details.group(3):
        print('=> The 2.4Ghz Channel Width is:', two_details.group(3))
    else:
        pass
    print('*---*-*---*-*---*-*---*-*---*-*---*')
    print("==> Getting 2.4Ghz clients associates with", ap_name)
    output1 = net_connect.send_command("sh wireless client ap dot11 24ghz chassis active r0 | inc" + " " + ap_name,
                                       read_timeout=603)
    print('MAC Address        AP Id             Status          WLAN Id  Authenticated')
    # filtering output for 2.4Ghz
    two = re.findall(r'.*', output1)
    two_new = [ele for ele in two if ele.strip()]
    for match in two_new:
        print(match)
    print('=> Total 2.4GHz clients connected =', len(two_new))
    print('*---*-*---*-*---*-*---*-*---*-*---*')
    total_client = int(len(two_new)) + int(len(five_new))
    print('==> Total Clients connected to' + " " + ap_name + " " + 'is' + " " + str(total_client))


def check_bgp_received_routes(bgp_nei, isbBgp_route, net_connect):
    if isbBgp_route:
        print('==> Showing received routes for this network||IP: ', isbBgp_route)
        output = net_connect.send_command(
            'sh ip bgp neighbo' + " " + bgp_nei + " " + 'received-rou' + " " + '| inc ^BGP|^Status|RIB-fa|best'
                                                                               '-external|secondary '
                                                                               'path|^Origin|^RPKI|' + " " +
            isbBgp_route)
        print(output)
    else:
        print('==> Showing all received routes')
        output = net_connect.send_command(
            'sh ip bgp neighbo' + " " + bgp_nei + " " + 'received-rou')
        print(output)

    return


def check_bgp_advertised_routes():
    pass


def check_wlc_snmp(ip, community):
    if '10.5.160.10' in ip or '10.5.144.10' in ip:
        COMMUNITY_STRING = 'Ic3L@nD'
        SNMP_PORT = 161
        a_device = (ip, community, SNMP_PORT)
        print('==> Date =', get_time_date()[0], '=> Time =', get_time_date()[1])
        try:
            if snmp_get_oid(a_device, oid='.1.3.6.1.2.1.1.5.0', display_errors=True):
                name = snmp_get_oid(a_device, oid='.1.3.6.1.2.1.1.5.0', display_errors=True)
                name_output = snmp_extract(name)
                print('==> WLC name =', name_output)

            if snmp_get_oid(a_device, oid='1.3.6.1.4.1.9.9.618.1.8.4.0', display_errors=True):
                total_ap = snmp_get_oid(a_device, oid='1.3.6.1.4.1.9.9.618.1.8.4.0', display_errors=True)
                total_output = snmp_extract(total_ap)
                print('==> Total AP joined WLC =', total_output)

            if snmp_get_oid(a_device, oid='1.3.6.1.4.1.9.9.618.1.8.12.0', display_errors=True):
                total_clients = snmp_get_oid(a_device, oid='1.3.6.1.4.1.9.9.618.1.8.12.0', display_errors=True)
                clients_output = snmp_extract(total_clients)
                print('==> Total Clients Connected to WLC =', clients_output)

            if 'PLC-PEV2' in name_output:
                pev2_mednet = '1'
                pev2_crewnet = '10'
                pev2_crew_compass = '3'
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.1', display_errors=True):
                    mednet_pev2 = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.1', display_errors=True)
                    mednet_pev2_output = snmp_extract(mednet_pev2)
                    print('==> Total Clients connected to MedNet =', mednet_pev2_output)
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.3', display_errors=True):
                    crew_compass_pev2 = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.3',
                                                     display_errors=True)
                    crew_compass_pev2_output = snmp_extract(crew_compass_pev2)
                    print('==> Total Clients connected to CrewCompass =', crew_compass_pev2_output)

            if 'PCL-PRC' in name_output:
                prc_mednet = '7'
                prc_crewnet = '8'
                prc_crew_compass = '4'
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.7', display_errors=True):
                    mednet_prc = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.7', display_errors=True)
                    mednet_prc_output = snmp_extract(mednet_prc)
                    print('==> Total Clients connected to MedNet =', mednet_prc_output)
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.8', display_errors=True):
                    crewnet_prc = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.8', display_errors=True)
                    crewnet_prc_output = snmp_extract(crewnet_prc)
                    print('==> Total Clients connected to CrewNet =', crewnet_prc_output)
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.4', display_errors=True):
                    crew_compass_prc = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.4', display_errors=True)
                    crew_compass_prc_output = snmp_extract(crew_compass_prc)
                    print('==> Total Clients connected to CrewCompass =', crew_compass_prc_output)

        # except TypeError:
        except NameError:
            pass
        print('*---*-*---*-*---*-*---*')

    else:
        # COMMUNITY_STRING = 'msdp725'
        SNMP_PORT = 161
        a_device = (ip, community, SNMP_PORT)
        print('==> Date =', get_time_date()[0], '=> Time =', get_time_date()[1])
        try:
            if snmp_get_oid(a_device, oid='.1.3.6.1.2.1.1.5.0', display_errors=True):
                name = snmp_get_oid(a_device, oid='.1.3.6.1.2.1.1.5.0', display_errors=True)
                name_output = snmp_extract(name)
                print('==> WLC name =', name_output)

            if snmp_get_oid(a_device, oid='1.3.6.1.4.1.9.9.618.1.8.4.0', display_errors=True):
                total_ap = snmp_get_oid(a_device, oid='1.3.6.1.4.1.9.9.618.1.8.4.0', display_errors=True)
                total_output = snmp_extract(total_ap)
                print('==> Total AP joined WLC =', total_output)

            if snmp_get_oid(a_device, oid='1.3.6.1.4.1.9.9.618.1.8.12.0', display_errors=True):
                total_clients = snmp_get_oid(a_device, oid='1.3.6.1.4.1.9.9.618.1.8.12.0', display_errors=True)
                clients_output = snmp_extract(total_clients)
                print('==> Total Clients Connected to WLC =', clients_output)

            if 'PCLAP' in name_output:
                ap_mednet = '5'
                ap_crewnet = '9'
                ap_crew_compass = '6'
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.5', display_errors=True):
                    mednet_ap = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.5', display_errors=True)
                    mednet_ap_output = snmp_extract(mednet_ap)
                    print('==> Total Clients connected to MedNet =', mednet_ap_output)
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.9', display_errors=True):
                    crewnet_ap = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.9', display_errors=True)
                    crewnet_ap_output = snmp_extract(crewnet_ap)
                    print('==> Total Clients connected to CrewNet =', crewnet_ap_output)
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.6', display_errors=True):
                    crew_compass_ap = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.6', display_errors=True)
                    crew_compass_ap_output = snmp_extract(crew_compass_ap)
                    print('==> Total Clients connected to CrewCompass =', crew_compass_ap_output)

            if 'PCLCB' in name_output:
                cb_mednet = '13'
                cb_crewnet = '14'
                cb_crew_compass = '16'
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.13', display_errors=True):
                    mednet_cb = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.13', display_errors=True)
                    mednet_cb_output = snmp_extract(mednet_cb)
                    print('==> Total Clients connected to MedNet =', mednet_cb_output)
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.14', display_errors=True):
                    crewnet_cb = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.14', display_errors=True)
                    crewnet_cb_output = snmp_extract(crewnet_cb)
                    print('==> Total Clients connected to CrewNet =', crewnet_cb_output)
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.16', display_errors=True):
                    crew_compass_cb = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.16', display_errors=True)
                    crew_compass_cb_output = snmp_extract(crew_compass_cb)
                    print('==> Total Clients connected to CrewCompass =', crew_compass_cb_output)

            if 'PCLCO' in name_output:
                cb_mednet = '11'
                cb_crewnet = '12'
                cb_crew_compass = '14'
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.11', display_errors=True):
                    mednet_co = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.11', display_errors=True)
                    mednet_co_output = snmp_extract(mednet_co)
                    print('==> Total Clients connected to MedNet =', mednet_co_output)
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.12', display_errors=True):
                    crewnet_co = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.12', display_errors=True)
                    crewnet_co_output = snmp_extract(crewnet_co)
                    print('==> Total Clients connected to CrewNet =', crewnet_co_output)
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.14', display_errors=True):
                    crew_compass_co = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.14', display_errors=True)
                    crew_compass_co_output = snmp_extract(crew_compass_co)
                    print('==> Total Clients connected to CrewCompass =', crew_compass_co_output)

            if 'PCLDI' in name_output:
                cb_mednet = '13'
                cb_crewnet = '14'
                cb_crew_compass = '9'
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.13', display_errors=True):
                    mednet_di = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.13', display_errors=True)
                    mednet_di_output = snmp_extract(mednet_di)
                    print('==> Total Clients connected to MedNet =', mednet_di_output)
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.14', display_errors=True):
                    crewnet_di = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.14', display_errors=True)
                    crewnet_di_output = snmp_extract(crewnet_di)
                    print('==> Total Clients connected to CrewNet =', crewnet_di_output)
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.9', display_errors=True):
                    crew_compass_di = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.9', display_errors=True)
                    crew_compass_di_output = snmp_extract(crew_compass_di)
                    print('==> Total Clients connected to CrewCompass =', crew_compass_di_output)

            if 'PCLEP' in name_output:
                cb_mednet = '12'
                cb_crewnet = '14'
                cb_crew_compass = '8'
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.12', display_errors=True):
                    mednet_ep = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.12', display_errors=True)
                    mednet_ep_output = snmp_extract(mednet_ep)
                    print('==> Total Clients connected to MedNet =', mednet_ep_output)
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.14', display_errors=True):
                    crewnet_ep = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.14', display_errors=True)
                    crewnet_ep_output = snmp_extract(crewnet_ep)
                    print('==> Total Clients connected to CrewNet =', crewnet_ep_output)
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.8', display_errors=True):
                    crew_compass_ep = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.8', display_errors=True)
                    crew_compass_ep_output = snmp_extract(crew_compass_ep)
                    print('==> Total Clients connected to CrewCompass =', crew_compass_ep_output)

            if 'PCLGP' in name_output:
                gp_mednet = '14'
                gp_crewnet = '15'
                gp_crew_compass = '19'
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.14', display_errors=True):
                    mednet_gp = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.14', display_errors=True)
                    mednet_gp_output = snmp_extract(mednet_gp)
                    print('==> Total Clients connected to MedNet =', mednet_gp_output)
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.15', display_errors=True):
                    crewnet_gp = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.15', display_errors=True)
                    crewnet_gp_output = snmp_extract(crewnet_gp)
                    print('==> Total Clients connected to CrewNet =', crewnet_gp_output)
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.19', display_errors=True):
                    crew_compass_gp = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.19', display_errors=True)
                    crew_compass_gp_output = snmp_extract(crew_compass_gp)
                    print('==> Total Clients connected to CrewCompass =', crew_compass_gp_output)

            if 'PCLIP' in name_output:
                ip_mednet = '9'
                ip_crewnet = '12'
                ip_crew_compass = '16'
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.9', display_errors=True):
                    mednet_ip = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.9', display_errors=True)
                    mednet_ip_output = snmp_extract(mednet_ip)
                    print('==> Total Clients connected to MedNet =', mednet_ip_output)
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.12', display_errors=True):
                    crewnet_ip = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.12', display_errors=True)
                    crewnet_ip_output = snmp_extract(crewnet_ip)
                    print('==> Total Clients connected to CrewNet =', crewnet_ip_output)
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.16', display_errors=True):
                    crew_compass_ip = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.16', display_errors=True)
                    crew_compass_ip_output = snmp_extract(crew_compass_ip)
                    print('==> Total Clients connected to CrewCompass =', crew_compass_ip_output)

            if 'PCLKP' in name_output:
                kp_mednet = '14'
                kp_crewnet = '15'
                kp_crew_compass = '13'
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.14', display_errors=True):
                    mednet_kp = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.14', display_errors=True)
                    mednet_kp_output = snmp_extract(mednet_kp)
                    print('==> Total Clients connected to MedNet =', mednet_kp_output)
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.15', display_errors=True):
                    crewnet_kp = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.15', display_errors=True)
                    crewnet_kp_output = snmp_extract(crewnet_kp)
                    print('==> Total Clients connected to CrewNet =', crewnet_kp_output)
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.13', display_errors=True):
                    crew_compass_kp = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.13', display_errors=True)
                    crew_compass_kp_output = snmp_extract(crew_compass_kp)
                    print('==> Total Clients connected to CrewCompass =', crew_compass_kp_output)

            if 'PCLMJ' in name_output:
                mj_mednet = '9'
                mj_crewnet = '10'
                mj_crew_compass = '11'
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.9', display_errors=True):
                    mednet_mj = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.9', display_errors=True)
                    mednet_mj_output = snmp_extract(mednet_mj)
                    print('==> Total Clients connected to MedNet =', mednet_mj_output)
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.10', display_errors=True):
                    crewnet_mj = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.10', display_errors=True)
                    crewnet_mj_output = snmp_extract(crewnet_mj)
                    print('==> Total Clients connected to CrewNet =', crewnet_mj_output)
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.10', display_errors=True):
                    crew_compass_mj = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.10', display_errors=True)
                    crew_compass_mj_output = snmp_extract(crew_compass_mj)
                    print('==> Total Clients connected to CrewCompass =', crew_compass_mj_output)

            if 'PCLRP' in name_output:
                rp_mednet = '8'
                rp_crewnet = '9'
                rp_crew_compass = '14'
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.8', display_errors=True):
                    mednet_rp = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.8', display_errors=True)
                    mednet_rp_output = snmp_extract(mednet_rp)
                    print('==> Total Clients connected to MedNet =', mednet_rp_output)
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.9', display_errors=True):
                    crewnet_rp = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.9', display_errors=True)
                    crewnet_rp_output = snmp_extract(crewnet_rp)
                    print('==> Total Clients connected to CrewNet =', crewnet_rp_output)
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.14', display_errors=True):
                    crew_compass_rp = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.14', display_errors=True)
                    crew_compass_rp_output = snmp_extract(crew_compass_rp)
                    print('==> Total Clients connected to CrewCompass =', crew_compass_rp_output)

            if 'PCLRU' in name_output:
                ru_mednet = '16'
                ru_crewnet = '8'
                ru_crew_compass = '14'
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.16', display_errors=True):
                    mednet_ru = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.16', display_errors=True)
                    mednet_ru_output = snmp_extract(mednet_ru)
                    print('==> Total Clients connected to MedNet =', mednet_ru_output)
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.8', display_errors=True):
                    crewnet_ru = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.8', display_errors=True)
                    crewnet_ru_output = snmp_extract(crewnet_ru)
                    print('==> Total Clients connected to CrewNet =', crewnet_ru_output)
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.14', display_errors=True):
                    crew_compass_ru = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.14', display_errors=True)
                    crew_compass_ru_output = snmp_extract(crew_compass_ru)
                    print('==> Total Clients connected to CrewCompass =', crew_compass_ru_output)

            if 'PCLSA' in name_output:
                ru_mednet = '12'
                ru_crewnet = '13'
                ru_crew_compass = '10'
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.12', display_errors=True):
                    mednet_sa = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.12', display_errors=True)
                    mednet_sa_output = snmp_extract(mednet_sa)
                    print('==> Total Clients connected to MedNet =', mednet_sa_output)
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.13', display_errors=True):
                    crewnet_sa = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.13', display_errors=True)
                    crewnet_sa_output = snmp_extract(crewnet_sa)
                    print('==> Total Clients connected to CrewNet =', crewnet_sa_output)
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.10', display_errors=True):
                    crew_compass_sa = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.10', display_errors=True)
                    crew_compass_sa_output = snmp_extract(crew_compass_sa)
                    print('==> Total Clients connected to CrewCompass =', crew_compass_sa_output)

            if 'PCLYP' in name_output:
                ru_mednet = '4'
                ru_crewnet = '5'
                ru_crew_compass = '6'
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.4', display_errors=True):
                    mednet_yp = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.4', display_errors=True)
                    mednet_yp_output = snmp_extract(mednet_yp)
                    print('==> Total Clients connected to MedNet =', mednet_yp_output)
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.5', display_errors=True):
                    crewnet_yp = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.5', display_errors=True)
                    crewnet_yp_output = snmp_extract(crewnet_yp)
                    print('==> Total Clients connected to CrewNet =', crewnet_yp_output)
                if snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.6', display_errors=True):
                    crew_compass_yp = snmp_get_oid(a_device, oid='1.3.6.1.4.1.14179.2.1.1.1.38.6', display_errors=True)
                    crew_compass_yp_output = snmp_extract(crew_compass_yp)
                    print('==> Total Clients connected to CrewCompass =', crew_compass_yp_output)

        # except TypeError:
        except NameError:
            pass
        print('*---*-*---*-*---*-*---*')


def check_power_inline_details(inter, net_connect):
    # getting and validating data
    consumption = net_connect.send_command('sh power inline' + " " + inter + " " + 'de' + " " + '| inc Measured')
    if '0.0' in consumption:
        print('==> No power consumption on the port')
    else:
        # getting data
        output = net_connect.send_command('sh power inline' + " " + inter + " " + 'de')

        # filtering details
        interface = re.search(r'Inter.*', output)
        inline = re.search(r'Inline.*', output)
        operation = re.search(r'Opera.*', output)
        device_detected = re.search(r'Device De.*', output)
        device_type = re.search(r'Device Ty.*', output)
        power_allocated = re.search(r'Power All.*', output)
        admin_value = re.search(r'Admin.*', output)
        power_drawn_from = re.search(r'Power drawn f.*', output)
        power_available = re.search(r'Power avai.*', output)
        actual_consumption = re.search(r'Actu.*', output)
        measured = re.search(r'Meas.*', output)
        max_power = re.search(r'Maxi.*', output)
        print('\n')

        # printing results:
        print(interface.group())
        print(inline.group())
        print(operation.group())
        print(device_detected.group())
        print(device_type.group())
        print('\n')
        print(power_allocated.group())
        print(admin_value.group())
        print(power_drawn_from.group())
        print(power_available.group())
        print('\n')
        print(actual_consumption.group())
        print(measured.group())
        print(max_power.group())

    return


def get_time_date():
    date_time = datetime.datetime.now().strftime("%H:%M:%S")
    date = datetime.date.today()
    # print('Current time = %s' % time)

    return date, date_time


def get_device_date_time(net_connect):
    get_time = net_connect.send_command('show clock | exc NTP')

    # filtering results
    filtered_time = re.search(r'(\S+)\s(\w+)\s(.*)', get_time).group(1)
    time_new = filtered_time.lstrip('*')
    filtered_timezone = re.search(r'(\S+)\s(\w+)\s(.*)', get_time).group(2)
    filtered_date = re.search(r'(\S+)\s(\w+)\s(.*)', get_time).group(3)
    print('=> Local time is:', time_new + " " + filtered_timezone + " " + '=> Date is:', filtered_date)


def create_folder_logs():
    # path for folder creation
    path = os.path.join("C:\\", "Jose", "python_ouput", str(get_time_date()[0]))

    # getting today date
    print("Today's date is: %s " % get_time_date()[1])
    isExist = os.path.exists(path)

    # validating if the folder exist
    if isExist:
        pass
    else:
        os.makedirs(path)


def panos_show_system_info(net_connect):
    output = net_connect.send_command('show system info')
    device_time = net_connect.send_command('show clock')
    device_time_filteres = re.search(r'\w.*', device_time)

    if not output:
        print("No values to show")
    else:
        # filtering the timne day + hours
        detail_time = re.search(r'upti\w+.\s*(\d+)\s\S+\s(.*)', output)
        name_detail = re.search(r'host\S+\s(.*)', output).group(1)
        print('\n')
        print('=====>*** HSC, Copy and paste the following output into the ticket <=====***\n')

        print('==> Getting the time and date<==')
        print('=> (EST) Date =', get_time_date()[0], '=> (EST) Time =', get_time_date()[1])
        print('=> Device local time and date is: ' + device_time_filteres.group() + '\n')

        print(
            name_detail + ' has been up and rechable since ' + detail_time.group(1) + ' days and ' + detail_time.group(
                2) + ' hours')

        # hostname
        hostname = re.search(r'host.*', output)
        print(hostname.group())

        # ip-address
        ipv4 = re.search(r'ip-add.*', output)
        print(ipv4.group())

        # netmask
        netmask = re.search(r'netm.*', output)
        print(netmask.group())

        # defaul-gw
        dw = re.search(r'defa.*', output)
        print(dw.group())

        # ip-assignment
        ipv4_assignment = re.search(r'ip-ass.*', output)
        print(ipv4_assignment.group())

        # mac-add
        mac_add = re.search(r'mac-ad.*', output)
        print(mac_add.group())

        # time
        time = re.search(r'time.*', output)
        print(time.group())

        # uptime
        uptime = re.search(r'upt.*', output)
        print(uptime.group())

        # family
        family = re.search(r'fam.*', output)
        print(family.group())

        # model
        model = re.search(r'mode.*', output)
        print(model.group())

        # serial
        serial = re.search(r'ser\S+\s\d{8}(\d+)', output)
        print('serial: XXXXXXXX' + serial.group(1) + ' # Printed the last 4 numbers')

        # software_version
        sw_version = re.search(r'sw-v.*', output)
        print(sw_version.group())


def panos_ha_state(net_connect):
    output = net_connect.send_command('show high-availability state')
    ha = re.search(r'Group\s1.*[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)',
                   output)
    ha_state = re.search(r'Sta\S+\s(\w+)', ha.group(5))
    ha_mode = re.search(r'Mod\S+\s(.*)', ha.group(1))
    # print('=> High Availability Mode =', ha_mode.group(1))
    # print('=> High Availability State =', ha_state.group(1))

    return ha_mode.group(1), ha_state.group(1)


def panos_check_interface(net_connect):
    # getting interface
    interface = str(input("Whats the interface:"))
    sysinfo = net_connect.send_command('show system info')
    output = net_connect.send_command('show interface' + " " + interface)
    ha = panos_ha_state(net_connect)
    if 'passive' in ha[1]:
        print('==>**** This is a STANDBY device, a lot of interfaces may be down ****<==')

    hostname = re.search(r'host\w+.\s(.*)', sysinfo)
    ipv4 = re.search(r'ip-ad\S+\s(\d+.*)', sysinfo)
    uptime = re.search(r'upt\S+\s(.*)', sysinfo)
    model = re.search(r'mod\S+\s(P.*)', sysinfo)
    print('*---*-*---*-*---*-*---*-')
    print('=> Hostname =', hostname.group(1))
    print('=> IP =', ipv4.group(1))
    print('=> UPTIME =', uptime.group(1))
    print('=> Model =', model.group(1))
    print('=> High Availability Mode =', ha[0])
    print('=> High Availability State =', ha[1])
    print('*---*-*---*-*---*-*---*-')
    print('=> Interface =', interface)
    link_data = re.search(r'Lin.*[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)[\r\n]+([^\r\n]+)',
                          output)
    if 'down' in output:
        print('=> The interface is down')
    else:
        try:
            if link_data.group():
                print(link_data.group())
        except AttributeError:
            pass
        print('*---*-*---*-*---*-*---*-')
        vr = re.search(r'Vir.*', output)
        mtu = re.search(r'.* \bMTU\b. *', output)
        inter_ip = re.search(r'Interface\sI.*', output)
        inter_mgmt = re.search(r'Interface\sma.*', output)
        zone = re.search(r'Zon.*', output)
        try:
            if vr.group():
                print(vr.group())
            if mtu.group():
                print(mtu.group())
            if inter_ip.group():
                print(inter_ip.group())
            if inter_mgmt.group():
                print(inter_mgmt.group())
            if zone.group():
                print(zone.group())
        except AttributeError:
            pass
        print('*---*-*---*-*---*-*---*-')


def panos_filter_logs(src_ip, net_connect):
    output = net_connect.send_command('show log traffic src in' + " " + src_ip + " " + 'receive_time in last-hour')
    print(output)


if __name__ == '__main__':
    inter = []
    vlan_id = input("Whats vlan#: ")

    JC, inter = get_credentials_and_interface()
    net_connect = ConnectHandler(**JC)
    net_connect.enable()

    inter_split = inter.split()
